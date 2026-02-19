# ======================================================
# SISTEMA ALMACEN - DESPACHO LA PAZ
# JC CODE SISTEM - VERSION CENTRADA PROFESIONAL
# ======================================================

import streamlit as st
import warnings
import plotly.express as px
import pandas as pd
import os
import altair as alt
import plotly.graph_objects as go
import xlsxwriter
from pathlib import Path
import glob
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime
from pyfiglet import Figlet


warnings.filterwarnings("ignore")

# ==============================
# CONFIGURACI√ìN
# ==============================
st.set_page_config(
    page_title="Sistema Almacen - Despacho La Paz",
    page_icon="üì¶",
    #layout="centered",
    layout="wide",
    initial_sidebar_state="expanded"  # üëà SIDEBAR VISIBLE
)

# ==============================
# CSS PROFESIONAL
# ==============================
st.markdown("""
<style>

/* ===== OCULTAR ELEMENTOS STREAMLIT ===== */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}

/* ===== FONDO GENERAL ===== */
.stApp {
    background: linear-gradient(135deg, #0e1117, #161b22);
}

/* ===== CONTENIDO CENTRADO ===== */
.block-container {
            
    max-width: 100%;
    padding-top: 1rem;
    padding-left: 2rem;
    padding-right: 2rem;
}

/* ===== LOGIN ===== */
.login-wrapper {
    height: 15vh;
    display: flex;
    justify-content: center;
    align-items: center;
}

.login-box {
    background-color: #161b22;
    padding: 40px;
    border-radius: 15px;
    width: 100%;
    max-width: 400px;
    box-shadow: 0px 0px 30px rgba(0,0,0,0.5);
}

.login-title {
    text-align: center;
    font-size: 24px;
    font-weight: bold;
    color: white;
    margin-bottom: 5px;
}

.login-sub {
    text-align: center;
    color: #8b949e;
    margin-bottom: 25px;
}

/* ===== INPUTS ===== */
div[data-testid="stTextInput"] {
    max-width: 600px;   /* üëà Cambia el tama√±o aqu√≠ */
    margin: auto;       /* üëà Lo centra */
}
.stTextInput>div>div>input {
    background-color: #0d1117;
    color: white;
    border-radius: 8px;
}

/* ===== BOTONES ===== */
div[data-testid="stButton"] {
    max-width: 250px;   /* üëà cambia el ancho aqu√≠ */
    margin-left: auto;
    margin-right: auto;
}
.stButton>button {
    background-color: #238636;
    color: white;
    border-radius: 8px;
    height: 40px;
    font-weight: bold;
    border: none;
}

.stButton>button:hover {
    background-color: #2ea043;
}

/* ===== SIDEBAR ===== */
section[data-testid="stSidebar"] {
    background-color: #161b22;
}

/* ===== RESPONSIVE ===== */
@media (max-width: 768px) {
    .block-container {
        padding-left: 1rem;
        padding-right: 1rem;
    }
}

</style>
""", unsafe_allow_html=True)

# ==============================
# USUARIOS Y PERMISOS
# ==============================
USUARIOS = {
    "admin": {"password": "Juan@666$.", "rol": "admin"},
    "despacho": {"password": "polysistemas$26", "rol": "despacho"},
    "despacho2": {"password": "poly$26.", "rol": "despacho2"}
}

PERMISOS = {
    "admin": [
        "PORTADA",
        "FILTRADOR ALMACEN",
        "CALENDARIO ALMACEN",
        "CREAR GUIA RUTA LA PAZ",
        "CHECK FILEWEB AND LASERFICHE",
        "SINTAXIS LASERFICHE-ONEIL",
        "DUPLICAR FILAS",
        "JUEGA CON DINO"
    ],
    "despacho": [
        "PORTADA",
        "FILTRADOR ALMACEN"
        
    ],
    "despacho2": [
        "PORTADA",
        "FILTRADOR ALMACEN",
        "CALENDARIO ALMACEN"
    ]
}

# ==============================
# SESSION STATE
# ==============================
if "login" not in st.session_state:
    st.session_state.login = False

# ==============================
# LOGIN
# ==============================
if not st.session_state.login:

    st.markdown("<div class='login-wrapper'>", unsafe_allow_html=True)
    #st.markdown("<div class='login-box'>", unsafe_allow_html=True)

    st.markdown("<div class='login-title'>üîê JC DEV_SISTEM</div>", unsafe_allow_html=True)
    st.markdown("<div class='login-sub'>SISTEMA ALMACEN DESPACHO LA PAZ</div>", unsafe_allow_html=True)

    usuario = st.text_input("üë§ Usuario")
    password = st.text_input("üîë Contrase√±a", type="password")

    if st.button("Ingresar", use_container_width=True):
        if usuario in USUARIOS and USUARIOS[usuario]["password"] == password:
            st.session_state.login = True
            st.session_state.usuario = usuario
            st.session_state.rol = USUARIOS[usuario]["rol"]
            st.rerun()
        else:
            st.error("‚ùå Usuario o contrase√±a incorrectos")

    st.markdown("</div></div>", unsafe_allow_html=True)
    st.stop()

# ==============================
# SIDEBAR
# ==============================
st.sidebar.markdown(
    "<h1 style='text-align:center; color:white;'>üì¶ JC CODE SISTEM</h1>",
    unsafe_allow_html=True
)

st.sidebar.write(f"üë§ Usuario: **{st.session_state.usuario}**")
st.sidebar.write(f"üîê Rol: **{st.session_state.rol}**")

if st.sidebar.button("Cerrar sesi√≥n"):
    st.session_state.clear()
    st.rerun()

st.sidebar.markdown("---")

opciones = PERMISOS[st.session_state.rol]
proyecto = st.sidebar.selectbox("Seleccione una opci√≥n", opciones)



# ==============================
# CONTENIDO PRINCIPAL
# ==============================
st.title("üìÅ SISTEMA ALMACEN LA PAZ - BOLIVIA")

if proyecto == "PORTADA":
   

    #------********************* IMAGEN PORTADA**********************************-----------

    # Crear objeto Figlet con fuente personalizada
    figlet = Figlet(font='standard')  # Puedes probar 'standard', 'big', 'banner3-D', etc.

    # Texto que quieres mostrar
    texto = "JC DEV SISTEM"

    # Convertir a ASCII art
    ascii_art = figlet.renderText(texto)

    # Mostrar en Streamlit centrado, con color ne√≥n y sombra
    st.markdown(f"""
    <div style="
        text-align: center;
        color: #00ffff;
        font-size: 24px;
        font-family: monospace;
        white-space: pre;
        text-shadow: 2px 2px 5px #00ffff, 0 0 10px #00ffff, 0 0 20px #00ffff;
    ">
    {ascii_art}
     """, 
    unsafe_allow_html=True)
   

    st.markdown("""
    <div style="width: 100%; overflow: hidden; white-space: nowrap;">
    <div style="
        display: inline-block;
        padding-left: 120%;
        animation: scroll-left 15s linear infinite;
        font-size: 18px;
        color: #3498db;
    ">
        DERECHOS-RECERVADOS - @JUAN CARLOS RAMOS CHURA - 2026
    </div>
    </div>

    <style>
    @keyframes scroll-left {
    0%   { transform: translateX(0%); }
    100% { transform: translateX(-100%); }
    }
    </style>
    """, unsafe_allow_html=True)


elif proyecto == "FILTRADOR ALMACEN":
    st.subheader("üì¶ FILTRADOR ALMACEN")
    st.markdown("<h1 style='text-align: center;'>FILTRADO DE DOCUMENTOS Y CAJAS ALMACEN</h1>", unsafe_allow_html=True)

    opciones = ["SELECCIONA UNA OPCION", "FILTRADO FILES", "FILTRADO TOMOS", "FILTRADO DE CAJAS"]

    # üîπ Creamos 3 columnas para centrar el contenido
    left, center, right = st.columns([1,2,1])
   
    with left:
        seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)

    if seleccion == "SELECCIONA UNA OPCION":   
        st.markdown("""
        <div style="
            text-align: center;
            padding: 10px;
            border-radius: 10px;
            background-color: #13161F;
            border: 1px solid #2d3748;
        ">
            <h1 style="color:white;">BIENVENIDO !! FILTRADO DE DOCUMENTOS ALMACEN</h1>
            <p style="color:#9ca3af;"><strong>Desarrollado por Juan Carlos Ramos Chura @2026</strong></p>
        </div>
        """, unsafe_allow_html=True)

    elif seleccion == "FILTRADO FILES":

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR FILES</h2>", unsafe_allow_html=True)

            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])


            if uploaded_file is not None:
            
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)
                
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
        
                st.write('FILTRADO POR NIVELES:')

                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES","Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])
                
                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":

                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                        file_path =  save_dir / f"{file_name}.xlsx"

                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )                           

                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")


                        save_dir = Path(ruta)

                        file_path =  save_dir / f"{file_name}.xlsx"

                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )        

                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")


                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
    
                # Mostrar un mensaje
                st.write('FILTRADO POR LOCACIONES:')

                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)
                        
                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            #st.write("Sube varios archivos Excel para combinarlos en uno solo.")
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar m√∫ltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Funci√≥n para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Files.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


            # ---------------------------------------------------------------------------------------------------------

    elif seleccion == "FILTRADO TOMOS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR TOMOS</h2>", unsafe_allow_html=True)

        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
                
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)
                        
                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "3":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "4":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "5":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "6":

                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'CAJA POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":
                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)

                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":

                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar m√∫ltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Funci√≥n para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Tomos.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")

            

            # ---------------------------------------------------------------------------------------------------------


    elif seleccion == "FILTRADO DE CAJAS":

        col1, col2 = st.columns(2)
        with col1:
            # Titulo de Aplicacion
        
            st.markdown("<h2 style='text-align: center;'>CARGAR PLANILLA DE EXCEL PARA FILTRAR CAJAS</h2>", unsafe_allow_html=True)
        
            #Cargar el archivo de excel 
            uploaded_file = st.file_uploader('Sube tu archivo de Excel', type=['xlsx','xls'])

        
            if uploaded_file is not None:
                # Leer el archivo Excel usando Pandas
                df = pd.read_excel(uploaded_file, engine='openpyxl')

                # Elimoinar Columnas
                Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8'], axis=1)
                Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
                Separar.columns = ['G','LA','P','S','N','L']
                Eliminar = pd.concat([Separar, Eliminar], axis=1)
                Eliminar = Eliminar.drop(['LOCACION'], axis=1)

                # Definimos una ruta para guardar nuestros archivos
                ruta = st.text_input("Introduce la ruta de la carpeta para guardar los Archivos: Obligatorio", "C:\\")
            
                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "Nivel_1", "Nivel_2", "Nivel_3", "Nivel_4", "Nivel_5", "Nivel_6"])

                Nivel = st.selectbox("Buscar Nivel", options = ["NIVEL", "1", "2", "3", "4", "5", "6"])

                if file_name == " " and Nivel == " ":
                    pass

                if Nivel == "1":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_1" and Nivel == "1":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        # Definir la ruta de guardado
                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "2":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_2" and Nivel == "2":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "3":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_3" and Nivel == "3":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Nivel == "4":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_4" and Nivel == "4":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "5":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_5" and Nivel == "5":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                if Nivel == "6":
                    Ordenar = Eliminar[(Eliminar['N'] == Nivel)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)

                    if file_name == "Nivel_6" and Nivel == "6":
                        
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


                # Permitir al usuario ingresar el nombre del archivo
                file_name = st.selectbox("Guardar Como:", options = ["OPCIONES", "L-DEV-CJ-001", "L-PREDESP_IN", "L-PREDESP_EX", "L-ING-CJ-001", "L-INV-CJ-001", "L-SCN-CJ-001", "L-DIG-CJ-001", "L-PALLET"])
                # Filtrado por Locacion
                Loc = st.selectbox("Buscar Locacion", options = ["LOCACION", "DEV", "PREDESP_IN", "PREDESP_EX", "ING", "INV", "SCN", "DIG", "PALLET"])

                if file_name == " " and Loc == " ":
                    pass

                if Loc == "DEV":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DEV-CJ-001" and Loc == "DEV":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "PREDESP_IN":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_IN" and Loc == "PREDESP_IN":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")
                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "PREDESP_EX":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PREDESP_EX" and Loc == "PREDESP_EX":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "ING":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-ING-CJ-001" and Loc == "ING":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "INV":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-INV-CJ-001" and Loc == "INV":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "SCN":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-SCN-CJ-001" and Loc == "SCN":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                if Loc == "DIG":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-DIG-CJ-001" and Loc == "DIG":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                        
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                
                if Loc == "PALLET":

                    Ordenar = Eliminar[(Eliminar['LA'] == Loc)]
                    Ordenar = Ordenar.sort_values(by=['G','LA', 'P', 'S', 'L', 'COD. POLY'])

                    st.dataframe(Ordenar)
                    
                    if file_name == "L-PALLET" and Loc == "PALLET":
                        # Limpiar el nombre del archivo para evitar caracteres problem√°ticos
                        file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    
                        save_dir = Path(ruta)

                        # Guardar el DataFrame en un archivo Excel f√≠sico
                        file_path =  save_dir / f"{file_name}.xlsx"

                        # Guardar el dataFrame en el archivo Excel
                        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                            Ordenar.to_excel(writer, index=False) # Guardar el primer dataFrame 

                        # Informar al usuario donde se guardo ek archivo
                        st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                        # Mostrar el boton de descarga para descargar el archivo guardado
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="Descargar Excel",
                                data=file,
                                file_name=f"{file_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )


            else:
                st.write("Por favor, suba un archivo de Excel para visualizarlo.")

        with col2:
            # Instrucciones       
            st.markdown("<h2 style='text-align: center;'>SUBE LOS ARCHIVOS FILTRADOS</h2>", unsafe_allow_html=True)

            # Cargar m√∫ltiples archivos
            uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)
                
                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Funci√≥n para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output

                # Convertir DataFrame combinado a Excel
                combined_file = convert_df_to_excel(combined_df)

                # Proporcionar el archivo combinado para descargar
                st.download_button(label="Descargar archivo Excel combinado",
                                data=combined_file,
                                file_name="Filtrado_Final_Cajas.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")

  


elif proyecto == "CALENDARIO ALMACEN":
    st.subheader("üìÖ CALENDARIO ALMACEN")
    import io
    
    st.markdown("<h1 style='text-align: center;'>CALENDARIO ALMACEN LA PAZ</h1>", unsafe_allow_html=True)

    st.markdown("<h2 style='text-align: center;'> CARGAR PLANILLA DE EXCEL PARA REALIZAR EL CALENDARIO</h2>", unsafe_allow_html=True)

    #Cargar el archivo de excel 
    uploaded_file = st.file_uploader('SUBE TU ARCHIVO DE EXCEL', type=['xlsx','xls'])

    if uploaded_file is not None:
    # Leer el archivo Excel usando Pandas
        df = pd.read_excel(uploaded_file, engine='openpyxl')

            # Elimoinar Columnas
        Eliminar = df.drop(['ELIMINAR_1', 'ELIMINAR_2', 'ELIMINAR_3', 'ELIMINAR_4', 'ELIMINAR_5','ELIMINAR_6','ELIMINAR_7','ELIMINAR_8','ELIMINAR_9','ELIMINAR_10','ELIMINAR_11','ELIMINAR_12','ELIMINAR_13'], axis=1)
        Separar = Eliminar["LOCACION"].str.split('[.-]', expand=True)
        Separar.columns = ['GALPON','LA','PASILLO','SHELF','NIVEL','COLUMNA']
        Eliminar = pd.concat([Separar, Eliminar], axis=1)
        Eliminar = Eliminar.drop(['LOCACION', 'LA'], axis=1)


        #st.dataframe(Eliminar)
        co1, co2 = st.columns(2)

        with co1:
            opciones = ["SELECCIONA UNA OPCION", "CALENDARIO GENERAL", "CALENDARIO POR PASILLO", "CALENDARIO POR SHELF"]

            seleccion = st.selectbox("Selecciona una opcion del menu: ", opciones)

        with co2:
            ruta = st.text_input("Introduce la ruta de la carpeta: ", "")

        # Funci√≥n para resaltar valores mayores a 10
        def highlight_integers(values):
            if values == 0:
                return 'background-color: green'
            elif values in range(1, 12):
                return 'background-color: orange'
            elif values in range(13, 25):
                return 'background-color: red'
            return ''
        

        if seleccion == "SELECCIONA UNA OPCION":   
            pass

        elif seleccion == "CALENDARIO GENERAL":

            Calendario = Eliminar.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
            
            # Reemplazar los valores None con 0
            Calendario = Calendario.fillna(0).astype(int)

            
            #----******************************************************---
            #SACAR DATOS DEL CALENDARIO 
            # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
            total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
            total_celdas = total_celdas.rename('TOTAL_SHELS')
            # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
            nonzero_mask = (Calendario != 0)
            total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
            total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
            # 3. ESPACIOS_LIBRES: suma de ceros
            zero_mask = (Calendario == 0)
            espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
            espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
            # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
            total_cajas_para_asignar = espacios_libres * 12
            total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
            # Unir todo en un solo DataFrame
            result = pd.concat([
                total_celdas,
                total_cajas_asignadas,
                espacios_libres,
                total_cajas_para_asignar
            ], axis=1).reset_index()
                
            # Contar cendas myores a 12 
            greater_than_12 = (Calendario > 12)
            # Contar celdas mayores a 12 por NIVEL
            count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

            # Resetear √≠ndice para mejor visualizaci√≥n
            count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
            count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

            #---*******************************************************---

            Calendario = Calendario.style.applymap(highlight_integers)
            
            st.write(Calendario)
            
            #----*****************DESCARGAR ARCHIVO*****************************

            from xlsxwriter.utility import xl_range

            # ‚úÖ Nombre seguro para archivo
            file_name = f"CALENDARIO_GENERAL_{pd.Timestamp.now().strftime('%Y-%m-%d_%H-%M-%S')}"
            safe_file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-") + ".xlsx"

            # ‚úÖ Quitar estilos de Calendario si los tiene
            calendario_clean = Calendario.data if hasattr(Calendario, 'data') else Calendario

            # ‚úÖ Crear archivo Excel en memoria
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book

                # üßæ HOJA: CALENDARIO COMPLETO
                sheetname = 'Calendario Completo'
                calendario_clean.to_excel(writer, sheet_name=sheetname, index=True)
                worksheet = writer.sheets[sheetname]

                # ‚ùÑÔ∏è Congelar primera fila y 3 primeras columnas
                worksheet.freeze_panes(2, 4)

                # üìê Determinar dimensiones
                n_rows, n_cols = calendario_clean.shape
                n_index_levels = calendario_clean.index.nlevels

                start_row = 1  # headers
                start_col = n_index_levels  # despu√©s de las columnas de √≠ndice

                end_row = start_row + n_rows - 1
                end_col = start_col + n_cols - 1

                cell_range = xl_range(start_row, start_col, end_row, end_col)

                # üé® Formatos condicionales
                formato_verde = workbook.add_format({'bg_color': '#92D050', 'border': 1})
                formato_naranja = workbook.add_format({'bg_color': '#FFC000', 'border': 1})
                formato_rojo = workbook.add_format({'bg_color': '#FF0000', 'border': 1})

                # ‚úÖ Condiciones de color
                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': '==',
                    'value': 0,
                    'format': formato_verde
                })

                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': 'between',
                    'minimum': 1,
                    'maximum': 11,
                    'format': formato_naranja
                })

                worksheet.conditional_format(cell_range, {
                    'type': 'cell',
                    'criteria': '>=',
                    'value': 12,
                    'format': formato_rojo
                })

                # üìè Ajuste autom√°tico de columnas
                for i, col in enumerate(calendario_clean.columns):
                    header = ' '.join(map(str, col)) if isinstance(col, tuple) else str(col)
                    max_len = max(len(header), 5)
                    worksheet.set_column(i + n_index_levels, i + n_index_levels, max_len + 2)

                # üîÑ Ajuste de ancho para columnas de √≠ndice
                for i in range(n_index_levels):
                    index_name = calendario_clean.index.names[i] or f"Index_{i}"
                    worksheet.set_column(i, i, len(index_name) + 4)

                # üßæ HOJAS ADICIONALES
                result.to_excel(writer, index=False, sheet_name='Datos por Galp√≥n')
                count_greater_than_12_by_nivel.to_excel(writer, index=False, sheet_name='Asignaciones en Exceso')
                Eliminar.to_excel(writer, index=False, sheet_name='Base de Datos')

            # üß∑ Finalizar el archivo
            output.seek(0)

            # üì• Bot√≥n de descarga en Streamlit
            st.download_button(
                label="üì• Descargar Excel con Formato",
                data=output.getvalue(),
                file_name=safe_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # üíæ Guardar localmente si se da ruta
            try:
                if ruta:
                    save_path = Path(ruta) / safe_file_name
                    with open(save_path, 'wb') as f:
                        f.write(output.getvalue())
                    st.success(f"‚úÖ Archivo guardado localmente en: {save_path.resolve()}")
            except Exception as e:
                st.warning(f"‚ö†Ô∏è No se pudo guardar localmente. Error: {str(e)}")

            
            # Mostrar en Streamlit

            colum1, colum2 = st.columns([12.04, 14.04])

            with colum1:
                #st.write("ASIGNACIONES QUE SOBREPASAN SU CANTIDAD LIMITE")
                st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)  
                st.table(count_greater_than_12_by_nivel)

            with colum2:
                #st.write("DATOS POR GAPON:")
                st.markdown("<h3 style='text-align: center;'>DATOS POR GAPON</h3>", unsafe_allow_html=True) 
                st.table(result)

            #---*******************************************************---

            #---*******************************************************---
            #  FILTRAR CODIGOS DE UN SHELF INDICADO
            #---*******************************************************---
            st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

            # Crear un selector para elegir una celda
            galpones = Calendario.index.get_level_values('GALPON').unique()
            pasillos = Calendario.index.get_level_values('PASILLO').unique()
            niveles = Calendario.index.get_level_values('NIVEL').unique()
            shelfs = Calendario.columns.levels[0].tolist()
            columnas = Calendario.columns.levels[1].tolist()
            
            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
            with col2:
                selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
            with col3:
                selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
            with col4:
                selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
            with col5:
                selected_columna = st.selectbox("Selecciona una Columna:", columnas)
            # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
            if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                filtered_data = Eliminar[
                    (Eliminar['GALPON'] == selected_galpon) &
                    (Eliminar['PASILLO'] == selected_pasillo) &
                    (Eliminar['NIVEL'] == selected_nivel) &
                    (Eliminar['SHELF'] == selected_shelf) &
                    (Eliminar['COLUMNA'] == selected_columna)
                ]

                # Asegurarse de que solo se muestren las columnas deseadas
                columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                # Renumerar los registros comensando desde 1
                filtered_data.index = filtered_data.index +1

                # Mostrar los c√≥digos poly en una tabla
                st.write("C√≥digos Poly asignados a la celda seleccionada:")
                st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

            #---********************************************************---

        elif seleccion == "CALENDARIO POR PASILLO":

            file_name = st.selectbox("Guardar Como:", options = ["OPCIONES","PASILLO_A", "PASILLO_B", "PASILLO_C", "PASILLO_D", "PASILLO_E", "PASILLO_F", "PASILLO_G", "PASILLO_H", "PASILLO_I"])
                        
            Pasillo = st.selectbox("Buscar Pasillo", options = ["PASILLO", "A", "B", "C", "D", "E", "F", "G", "H", "I"])

            if Pasillo == "PASILLO":
                pass

            if Pasillo == "A":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_A" and Pasillo == "A":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.write("DATOS POR GAPON:")
                    st.table(result)

                #---*******************************************************---              

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---
    
            if Pasillo == "B":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                #---********************************************************---
    
                if file_name == "PASILLO_B" and Pasillo == "B":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.write("DATOS POR GAPON:")
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

            #---*******************************************************************---            

            if Pasillo == "C":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')

                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)


                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_C" and Pasillo == "C":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "D":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_D" and Pasillo == "D":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "E":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_E" and Pasillo == "E":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATO POR GALPON</h3>", unsafe_allow_html=True) 
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "F":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                
                if file_name == "PASILLO_F" and Pasillo == "F":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---


            if Pasillo == "G":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_G" and Pasillo == "G":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            
            if Pasillo == "H":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---


                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)

                
                if file_name == "PASILLO_H" and Pasillo == "H":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---

            if Pasillo == "I":

                Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

                Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')
        
                # Reemplazar los valores None con 0
                Calendario = Calendario.fillna(0).astype(int)

                #----******************************************************---
                #SACAR DATOS DEL CALENDARIO 
                # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
                total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
                total_celdas = total_celdas.rename('TOTAL_SHELS')
                # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
                nonzero_mask = (Calendario != 0)
                total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
                total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
                # 3. ESPACIOS_LIBRES: suma de ceros
                zero_mask = (Calendario == 0)
                espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
                espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
                # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
                total_cajas_para_asignar = espacios_libres * 12
                total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
                # Unir todo en un solo DataFrame
                result = pd.concat([
                    total_celdas,
                    total_cajas_asignadas,
                    espacios_libres,
                    total_cajas_para_asignar
                ], axis=1).reset_index()
                
                # Contar cendas myores a 12 
                greater_than_12 = (Calendario > 12)
                # Contar celdas mayores a 12 por NIVEL
                count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

                # Resetear √≠ndice para mejor visualizaci√≥n
                count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
                count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

                #---*******************************************************---

                Calendario = Calendario.style.applymap(highlight_integers)

                st.write(Calendario)


                if file_name == "PASILLO_I" and Pasillo == "I":
                    
                    file_name = file_name.replace(" ", "_").replace(":", "-").replace("/", "-")

                    # Definir la ruta de guardado
                    save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos

                    #save_dir = Path(ruta)  # Directorio donde se guardar√°n los archivos


                    # Guardar el DataFrame en un archivo Excel f√≠sico
                    file_path =  save_dir / f"{file_name}.xlsx"

                    # Guardar el dataFrame en el archivo Excel
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        Calendario.to_excel(writer, index=True) # Guardar el primer dataFrame 

                    # Informar al usuario donde se guardo ek archivo
                    st.write(f"Archivo Guardo en: {file_path.resolve()} ")

                    # Mostrar el boton de descarga para descargar el archivo guardado
                    with open(file_path, "rb") as file:
                        st.download_button(
                        label="Descargar Excel",
                        data=file,
                        file_name=f"{file_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                #---********************************************************---
                # Mostrar en Streamlit
                colum1, colum2 = st.columns([12.04, 14.04])
                with colum1:
                    st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True)
                    st.table(count_greater_than_12_by_nivel)
                with colum2:
                    st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                    st.table(result)

                #---*******************************************************---

                #---*******************************************************---
                #  FILTRAR CODIGOS DE UN SHELF INDICADO
                #---*******************************************************---
                st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

                # Crear un selector para elegir una celda
                galpones = Calendario.index.get_level_values('GALPON').unique()
                pasillos = Calendario.index.get_level_values('PASILLO').unique()
                niveles = Calendario.index.get_level_values('NIVEL').unique()
                shelfs = Calendario.columns.levels[0].tolist()
                columnas = Calendario.columns.levels[1].tolist()
                
                col1, col2, col3, col4, col5 = st.columns(5)

                with col1:
                    selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
                with col2:
                    selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
                with col3:
                    selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
                with col4:
                    selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
                with col5:
                    selected_columna = st.selectbox("Selecciona una Columna:", columnas)
                # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
                if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                    filtered_data = Eliminar[
                        (Eliminar['GALPON'] == selected_galpon) &
                        (Eliminar['PASILLO'] == selected_pasillo) &
                        (Eliminar['NIVEL'] == selected_nivel) &
                        (Eliminar['SHELF'] == selected_shelf) &
                        (Eliminar['COLUMNA'] == selected_columna)
                    ]

                    # Asegurarse de que solo se muestren las columnas deseadas
                    columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                    filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                    # Renumerar los registros comensando desde 1
                    filtered_data.index = filtered_data.index +1

                    # Mostrar los c√≥digos poly en una tabla
                    st.write("C√≥digos Poly asignados a la celda seleccionada:")
                    st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---


        elif seleccion == "CALENDARIO POR SHELF":

            Pasillo = st.selectbox("Buscar Pasillo", options = ["PASILLO", "A", "B", "C", "D", "E", "F"])

            Calendario = Eliminar[(Eliminar["PASILLO"] == Pasillo)]

            Shelf = st.selectbox("Buscar Shelf", options = ["TODOS","01", "02", "03", "04", "05", "06", "07", "08", "09"]) 

         
           # Shelf = st.text_input("Ingrese Numero de Shelf: ")

            Calendario = Calendario[(Calendario["SHELF"] == Shelf)]

            # Crear una tabla dinamica para mostrar el calendario

            Calendario = Calendario.pivot_table(index=["GALPON", "PASILLO", "NIVEL"], columns=["SHELF", "COLUMNA"], values=['CODIGO POLY'], aggfunc='size')

            # Reemplazar los valores None con 0
            Calendario = Calendario.fillna(0).astype(int)

            #----******************************************************---
            #SACAR DATOS DEL CALENDARIO 
            # 1. TOTAL_CELDAS: filas por galp√≥n * columnas
            total_celdas = Calendario.groupby(level='GALPON').size() * Calendario.shape[1]
            total_celdas = total_celdas.rename('TOTAL_SHELS')
            # 2. TOTAL_CAJAS_ASIGNADAS: suma valores distintos de cero
            nonzero_mask = (Calendario != 0)
            total_cajas_asignadas = (nonzero_mask * Calendario).groupby(level='GALPON').sum().sum(axis=1)
            total_cajas_asignadas = total_cajas_asignadas.rename('TOTAL_CAJAS_ASIGNADAS')
            # 3. ESPACIOS_LIBRES: suma de ceros
            zero_mask = (Calendario == 0)
            espacios_libres = zero_mask.groupby(level='GALPON').sum().sum(axis=1)
            espacios_libres = espacios_libres.rename('ESPACIOS_LIBRES')
            # 4. TOTAL_CAJAS_PARA_ASIGNAR: multiplicacion de espacios libres por 12
            total_cajas_para_asignar = espacios_libres * 12
            total_cajas_para_asignar = total_cajas_para_asignar.rename('TOTAL_CAJAS_PARA_ASIGNAR')
            # Unir todo en un solo DataFrame
            result = pd.concat([
                total_celdas,
                total_cajas_asignadas,
                espacios_libres,
                total_cajas_para_asignar
            ], axis=1).reset_index()
                
            # Contar cendas myores a 12 
            greater_than_12 = (Calendario > 12)
            # Contar celdas mayores a 12 por NIVEL
            count_greater_than_12_by_nivel = greater_than_12.groupby(level=['GALPON','NIVEL']).sum().sum(axis=1)

            # Resetear √≠ndice para mejor visualizaci√≥n
            count_greater_than_12_by_nivel = count_greater_than_12_by_nivel.reset_index()
            count_greater_than_12_by_nivel.columns = ['GALPON', 'NIVEL', 'ASIGNACION EN EXCESO']

            #---*******************************************************---

            Calendario = Calendario.style.applymap(highlight_integers)

            st.write(Calendario)

            #---********************************************************---
            # Mostrar en Streamlit
            colum1, colum2 = st.columns([12.04, 14.04])
            with colum1:
                st.markdown("<h3 style='text-align: center;'>ASIGNACIONES EN EXCESO</h3>", unsafe_allow_html=True) 
                st.table(count_greater_than_12_by_nivel)
            with colum2:
                st.markdown("<h3 style='text-align: center;'>DATOS POR GALPON</h3>", unsafe_allow_html=True)
                st.table(result)

            #---*******************************************************---

            #---*******************************************************---
            #  FILTRAR CODIGOS DE UN SHELF INDICADO
            #---*******************************************************---
            st.markdown("<h3 style='text-align: center;'>MOSTRAR CAJAS DE SHEL SELECCIONADO</h3>", unsafe_allow_html=True)

            # Crear un selector para elegir una celda
            galpones = Calendario.index.get_level_values('GALPON').unique()
            pasillos = Calendario.index.get_level_values('PASILLO').unique()
            niveles = Calendario.index.get_level_values('NIVEL').unique()
            shelfs = Calendario.columns.levels[0].tolist()
            columnas = Calendario.columns.levels[1].tolist()
                
            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                selected_galpon = st.selectbox("Selecciona un Galp√≥n:", galpones)
            with col2:
                selected_pasillo = st.selectbox("Selecciona un Pasillo:", pasillos)
            with col3:
                selected_nivel = st.selectbox("Selecciona un Nivel:", niveles)
            with col4:
                selected_shelf = st.selectbox("Selecciona un Shelf:", shelfs)
            with col5:
                selected_columna = st.selectbox("Selecciona una Columna:", columnas)
            # Filtrar los c√≥digos poly seg√∫n la selecci√≥n
            if selected_galpon and selected_pasillo and selected_nivel and selected_shelf and selected_columna:
                filtered_data = Eliminar[
                    (Eliminar['GALPON'] == selected_galpon) &
                    (Eliminar['PASILLO'] == selected_pasillo) &
                    (Eliminar['NIVEL'] == selected_nivel) &
                    (Eliminar['SHELF'] == selected_shelf) &
                    (Eliminar['COLUMNA'] == selected_columna)
                ]

                # Asegurarse de que solo se muestren las columnas deseadas
                columns_to_show = ['GALPON', 'PASILLO', 'NIVEL','SHELF', 'COLUMNA', 'CODIGO POLY']
                filtered_data = filtered_data[columns_to_show].reset_index(drop=True)

                # Renumerar los registros comensando desde 1
                filtered_data.index = filtered_data.index +1

                # Mostrar los c√≥digos poly en una tabla
                st.write("C√≥digos Poly asignados a la celda seleccionada:")
                st.dataframe(filtered_data[['GALPON', 'PASILLO', 'NIVEL', 'SHELF', 'COLUMNA', 'CODIGO POLY']])

                #---********************************************************---


elif proyecto == "CREAR GUIA RUTA LA PAZ":
        import streamlit as st
        import pandas as pd
        import os
        from io import BytesIO

        st.subheader("üöö CREAR GUIA RUTA LA PAZ")

        # --------------------- SIDEBAR MENU ---------------------
        st.sidebar.title("Men√∫ de Opciones")
        menu = st.sidebar.radio("Ir a:", [
            "üè∑Ô∏è Filtrar Solicitudes",
            "üß© Combinar Archivos",
            "üîé Generar Sintaxis de B√∫squeda",
            "üì¶ Preparar Ruta Almac√©n La Paz",
            "üì¶ Preparar Guias de Ruta",
            "üóëÔ∏è Eliminar Archivos"

        ])
                                
        # --------------------- 1. FILTRAR SOLICITUDES V2 ---------------------
        def filtrar_solicitudes():

            st.title("üìä FILTRADOR DE SOLICITUDES PARA ELABORAR RUTAS")

            co1,co2 = st.columns([15.04, 8.05])
            with co1:
                # Subida del archivo
                uploaded_file = st.file_uploader("üìÅ CARGAR ARCHIVO EXCEL", type=["csv","xlsx"])
            #with co2:
                # Input para ruta personalizada (opcional)
                #carpeta_guardado = st.text_input(
                    #"üìÅ Ingresa la ruta para para guardar el Archivo Filtrado :",
                    #placeholder="Ej: C:/Users/TuUsuario/Desktop"
                #)

            # Definiciones de filtros
            TIPOS_SOLICITUD = [
                "SE1 - Env√≠o de Materiales",
                "SR1 - Recojo de items nuevos (Cajas o Files)",
                "SR2 - Recojo de Items por devoluci√≥n (Cajas o Files)"
            ]

            # Diccionario de Centros de Costo
            CENTROS_CENTRALIZADO = {
                "REGIONAL LA PAZ": {
                    "ZONA SUR": [
                        "212 - COTA COTA", 
                        "208 - SAN MIGUEL", 
                        "219 - OBRAJES"
                        ],
                    "ZONA ESTE": [
                        "204 - MIRAFLORES", 
                        "218 - VILLA ARMONIA", 
                        "211 - CRUCE VILLA COPACABANA", 
                        "223 - PAMPAHASI"],
                    "CENTRO": [
                        "221 - LA PORTADA", 
                        "205 - EL TEJAR", 
                        "216 - GRAN PODER", 
                        "202 - GARITA", "206 - ALONSO DE MENDOZA",
                        "213 - NORMALIZADORA LA PAZ", 
                        "214 - SOL AMIGO LA PAZ", 
                        "220 - TEMBLADERANI", 
                        "201 - SAN PEDRO",
                        "295 - OFICINA NACIONAL", 
                        "210 - CAMACHO", 
                        "209 - BALLIVIAN", 
                        "200 - REGIONAL LA PAZ"],
                    "NORTE": [
                        "203 - VILLA FATIMA", 
                        "224 - CHUQUIAGUILLO", 
                        "222 - PERIFERICA", 
                        "217 - VINO TINTO"]
                },
                "REGIONAL EL ALTO": {
                    "NORTE": [
                        "253 - RIO SECO", 
                        "276 - VILLA INGENIO", 
                        "272 - SAN ROQUE", 
                        "266 - MERCADO EL CARMEN R√çO SECO",
                        "275 - FRANZ TAMAYO", 
                        "251 - 16 DE JULIO", 
                        "279 - CHACALTAYA", 
                        "277 - FERROPETROL",
                        "280 - LAGUNAS EX PARADA 8"],
                    "CENTRO": [
                        "270 - 12 DE OCTUBRE", 
                        "252 - LA CEJA", 
                        "263 - VILLA DOLORES", 
                        "258 - NORMALIZACI√ìN EL ALTO",
                        "262 - SATELITE",
                        "264 - SOL AMIGO EL ALTO",
                        "265 - AGENCIA MOVIL",  
                        "250 - REGIONAL EL ALTO"],
                    "OESTE": [
                        "254 - VILLA ADELA", 
                        "261 - BOLIVIA", 
                        "274 - QUISWARAS"],
                    "SUR": [
                        "267 - SANTIAGO II", 
                        "273 - EL KENKO", 
                        "260 - SENKATA", 
                        "269 - VENTILLA"],
                    "VIACHA": [
                        "271 - INGAVI", 
                        "256 - VIACHA"]
                },
                "REGIONAL ORURO": {
                    "ORURO": [
                        "401 - CENTRAL", 
                        "407 - NORMALIZACION ORURO", 
                        "408 - PUNTO AMIGO ORURO", 
                        "409 - VIRGEN DEL SOCAV√ìN",
                        "410 - TAGARETE", 
                        "411 - TACNA"]
                },
                "REGIONAL SUCRE": {
                    "SUCRE": [
                        "100 - REGIONAL CHUQUISACA", 
                        "101 - MERCADO CAMPESINO SUCRE", 
                        "102 - NORMALIZACION SUCRE", 
                        "103 - ESPA√ëA",
                        "104 - SOL AMIGO SUCRE", 
                        "105 - GERMAN MENDOZA", 
                        "106 - CHARCAS", 
                        "107 - ZUDA√ëEZ",
                        "108 - LAS AMERICAS", 
                        "109 - LAJASTAMBO"]
                },
                "REGIONAL TARIJA": {
                    "": [
                        "601 - MERCADO CAMPESINO TARIJA",
                        "603 - SUR",
                        "604 - SOL AMIGO TARIJA",
                        "605 - 15 DE ABRIL",
                        "606 - NORMALIZACI√ìN TARIJA",
                        "607 - TABLADITA",
                        "608 - YACUIBA",
                        "609 - PALMARCITO",
                        "610 - MERCADO CAMPESINO YACUIBA"
                    ]
                },

                "BANCO NACIONAL DE BOLIVIA": {
                    "BNB - LA PAZ": [
                        "130 - BANCO NACIONAL DE BOLIVIA - LA PAZ"],
                    "BNB - COCHABAMBA":[
                        "132 - BANCO NACIONAL DE BOLIVIA - COCHABAMBA"],
                    "BNB - TARIJA":[
                        "136 - BANCO NACIONAL DE BOLIVIA - TARIJA"],
                    "BNB - ORURO":[
                        "134 - BANCO NACIONAL DE BOLIVIA - ORURO"],
                    "BNB - SUCRE":[
                        "133 - BANCO NACIONAL DE BOLIVIA - SUCRE"],
                    "BNB - SANTA CRUZ":[
                        "131 - BANCO NACIONAL DE BOLIVIA - SANTA CRUZ"],
                    "BNB - BENI":[
                        "137 - BANCO NACIONAL DE BOLIVIA - BENI"],
                    "BNB - POTOSI":[
                        "135 - BANCO NACIONAL DE BOLIVIA - POTOSI"],
                    "BNB - PANDO":[
                        "138 - BANCO NACIONAL DE BOLIVIA - PANDO"],
                    "BNB - POLYSISTEMAS":[
                        "C008 - POLYSISTEMAS"]

                },

                "BANCO FIE": {
                    "BFIE LA PAZ": [
                        "124-1 - BFIE LA PAZ"],
                    "BFIE SANTA CRUZ": [
                        "124-2 - BFIE SANTA CRUZ"],
                    "BFIE INTERNO": [   
                        "C008 - POLYSISTEMAS"]
                        
                },

                "JTI BOLIVIA": {
                    "JTI EXTERNO": [
                        "367 - JTI BOLIVIA"],
                    "JTI INTERNO": [   
                        "C008 - POLYSISTEMAS"]
                },

                "REGIONAL COCHABAMBA": {
                    "": [
                        "301 - ESTEBAN ARCE",
                        "302 - SAN MARTIN",
                        "303 - HUAYRA KHASA",
                        "305 - CRUCE TAQUI√ëA",
                        "306 - QUILLACOLLO",
                        "307 - COLCA PIRHUA",
                        "309 - MUYURINA",
                        "310 - NORMALIZACI√ìN COCHABAMBA",
                        "311 - RECAUDADORA JORDAN",
                        "312 - SACABA",
                        "313 - VILLA GALINDO",
                        "314 - PUNATA",
                        "316 - AYACUCHO",
                        "317 - SOL AMIGO COCHABAMBA",
                        "318 - PANAMERICANA",
                        "320 - CLIZA",
                        "321 - VINTO",
                        "322 - REP√öBLICA",
                        "323 - TIQUIPAYA",
                        "324 - QUINTANILLA",
                        "325 - JORDAN",
                        "326 - PLAZA BOLIVAR",
                        "327 - PETROLERA",
                        "328 - LA CHIMBA",
                        "329 - AMERICA",
                        "331 - EL AVION",
                        "332 - VILLA PAGADOR",
                        "333 - PACATA"
                    ]
                },
                "REGIONAL BENI": {
                    "": [
                        "801 - TRINIDAD",
                        "802 - RIBERALTA",
                        "803 - GUAYARAMERIN"
                    ]
                },
                "REGIONAL COBIJA": {
                    "": [
                        "901 - COBIJA",
                        "902 - TAJIBOS",
                        "903 - NORMALIZACI√ìN"
                    ]
                },
                "REGIONAL POTOSI": {
                    "": [
                        "501 - MERCADO UYUNI",
                        "502 - SOL AMIGO POTOSI",
                        "503 - BOULEVARD", 
                        "505 - LAS BANDERAS"
                    ]
                },
                "REGIONAL SANTA CRUZ": {
                    "": [
                        "701 - CASCO VIEJO",
                        "702 - EL PARI",
                        "703 - MUTUALISTA",
                        "704 - 1RO. DE MAYO",
                        "705 - MONTERO",
                        "706 - EL TORNO",
                        "709 - PIRAI",
                        "711 - PLAN 3000",
                        "715 - LA GUARDIA",
                        "716 - ALTO SAN PEDRO",
                        "718 - NORTE",
                        "719 - SOL AMIGO SANTA CRUZ",
                        "721 - ARROYO CONCEPCI√ìN",
                        "723 - PAMPA DE LA ISLA",
                        "724 - COLORADA",
                        "725 - Minero",
                        "726 - SAN JULIAN",
                        "727 - SAN JOSE",
                        "728 - 2 DE AGOSTO",
                        "730 - LOS LOTES",
                        "731 - GERMAN MORENO",
                        "732 - YAPACANI",
                        "734 - SATELITE NORTE",
                        "735 - NORTE I",
                        "736 - EL QUIOR",
                        "737 - VIRGEN DE LUJAN",
                        "738 - EL BAJIO",
                        "739 - CRISTO REDENTOR"
                    ]
                }
            }

            if uploaded_file:
                df = pd.read_excel(uploaded_file)
                
                col1, col2 = st.columns(2)

                with col1:
                    st.subheader("FILTRAR TIPO DE SOLICITUD")
                    tipos_seleccionados = st.multiselect(
                        "Selecciona uno o varios tipos de solicitud:",
                        options=TIPOS_SOLICITUD,
                        default=TIPOS_SOLICITUD
                    )
                with col2:
                    st.subheader("SELECCIONAR REGION Y SUBZONA")

                    region = st.selectbox("Selecciona una regi√≥n", options=list(CENTROS_CENTRALIZADO.keys()))
                    subzonas = list(CENTROS_CENTRALIZADO[region].keys())
                    subzona = st.selectbox("Selecciona una subzona", options=subzonas)

                    centros_disponibles = CENTROS_CENTRALIZADO[region][subzona]
                    centros_seleccionados = st.multiselect("Selecciona Centros de Costo", centros_disponibles, default=centros_disponibles)

                # Filtrado de datos
                df_filtrado = df[
                    (df["Tipo de Solicitud"].isin(tipos_seleccionados)) &
                    (df["Centro de Costo"].isin(centros_seleccionados))
                ]

                # Eliminar columnas no deseadas
                columnas_a_eliminar = ["Autorizado", "Locacion", "Centro de Costo Polysistemas", "Fecha de Impresi√≥n"]
                df_filtrado = df_filtrado.drop(columns=columnas_a_eliminar, errors='ignore')

                st.success(f"üîç Filtrado completo: {len(df_filtrado)} registros encontrados.")

                colum1, colum2 = st.columns([15.50, 6.10])

                with colum1:
                    # Mostrar tabla filtrada
                    st.dataframe(df_filtrado, use_container_width=True)

                    # Input para que el usuario elija el nombre del archivo
                    nombre_archivo = st.text_input(
                        "üìù Nombre del archivo a descargar:",
                        value="resultado_filtrado.xlsx",
                        placeholder="Ej: rutas_filtradas.xlsx"
                    )

                with colum2:
                    # Mostrar resumen basado en TipoFile y suma de Cantidad como tabla
                    tipos_resumen = ["CAJA", "Caja", "Cintillos", "FILE"]
                    st.subheader("üì¶ RESUMEN DE CANTIDADES")

                    if "TipoFile" in df_filtrado.columns and "Cantidad" in df_filtrado.columns:
                        resumen_data = []

                        for tipo in tipos_resumen:
                            total = df_filtrado[df_filtrado["TipoFile"] == tipo]["Cantidad"].sum()
                            resumen_data.append({"TipoFile": tipo, "Total Cantidad": int(total)})

                        resumen_df = pd.DataFrame(resumen_data)
                        st.table(resumen_df)
                    else:
                        st.warning("‚ö†Ô∏è Las columnas 'TipoFile' o 'Cantidad' no se encuentran en el archivo.")
                    
                    # -------------------------------
                    # BOT√ìN PARA DESCARGAR RESUMEN
                    # -------------------------------
                    resumen_output = BytesIO()
                    with pd.ExcelWriter(resumen_output, engine='openpyxl') as writer:
                        resumen_df.to_excel(writer, index=False, sheet_name='Resumen')
                    resumen_output.seek(0)

                    st.download_button(
                        label="üì• Descargar Resumen de Cantidades",
                        data=resumen_output,
                        file_name="resumen_cantidades.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    #else:
                    #    st.warning("‚ö†Ô∏è Las columnas 'TipoFile' o 'Cantidad' no se encuentran en el archivo.")


                # Descarga directa sin guardar en carpeta
                st.subheader("üì• Descargar Excel Filtrado")

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_filtrado.to_excel(writer, index=False, sheet_name='Filtrado')
                output.seek(0)

                st.download_button(
                    label="üì• Descargar Excel filtrado",
                    data=output,
                    #file_name="resultado_filtrado.xlsx",
                    file_name=nombre_archivo if nombre_archivo.strip() else "resultado_filtrado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


        # --------------------- 2. COMBINAR ARCHIVOS ---------------------
        def combinar_archivos():
            st.title("üìö COMBINAR ARCHIVOS FILTRADOS")
              

            co1,co2 = st.columns([15.04, 8.05])
            with co1:
                # Cargar m√∫ltiples archivos
                uploaded_files = st.file_uploader("Elige archivos Excel", type=["xlsx", "xls"], accept_multiple_files=True)
                
            #with co2:
                # Input para ruta personalizada (opcional)
                #carpeta_guardado = st.text_input(
                    #"üìÅ Ingresa la ruta para para guardar el Archivo Filtrado :",
                    #placeholder="Ej: C:/Users/TuUsuario/Desktop"
                #)

            # Comprobar si se han subido archivos
            if uploaded_files:

                # Ordenar los archivos por nombre, si es necesario
                uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

                dfs = []
                for file in uploaded_files:
                    # Leer cada archivo Excel en un DataFrame
                    df = pd.read_excel(file)
                    dfs.append(df)

                # Combinar todos los DataFrames en uno solo
                combined_df = pd.concat(dfs, ignore_index=True)

                # Mostrar el DataFrame combinado
                st.write("DataFrame Combinado:")
                st.dataframe(combined_df)

                # Funci√≥n para convertir el DataFrame combinado a Excel
                def convert_df_to_excel(df):
                    # Crear un objeto BytesIO
                    output = BytesIO()
                    # Escribir el DataFrame en el objeto BytesIO
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    # Mover el cursor al principio del objeto BytesIO
                    output.seek(0)
                    return output
                
                excel_data = convert_df_to_excel(combined_df)

                # Bot√≥n de descarga
                st.download_button(
                    label="üì• Descargar Excel Combinado",
                    data=excel_data,
                    file_name="Archivo_Combinado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            else:
                st.write("Por favor, sube los archivos Excel para combinarlos.")


        # --------------------- 3. GENERADOR DE SINTAXIS ---------------------
        def generar_sintaxis():

            st.title("üîé GENERADOR DE SINTAXIS DE B√öSQUEDA")

            col1, col2 = st.columns((2))
            with col1:
                # Cargar archivo Excel
                st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCode'</h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):

                    if uploaded_file is not None:
                        try:
                            # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                            df = pd.read_excel(uploaded_file)

                            # Verificar si la columna existe
                            if 'SolicitudCode' not in df.columns:
                                st.error("La columna 'SolicitudCode' no se encuentra en el archivo.")
                            else:
                                # Eliminar las dem√°s columnas y eliminar nulos
                                solicitud_codes = df['SolicitudCode'].dropna().astype(str).tolist()

                                # Crear la sintaxis
                                solicitud_sintaxis = " | ".join(
                                    [f'\n{{[Solicitud]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                )

                                sintaxis_final = f"{solicitud_sintaxis}"
                        except Exception as e:
                            st.error(f"Error al leer el archivo: {e}")

            with col2:
                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)

                    # Convertir a bytes para evitar errores de descarga
                    contenido_descarga = sintaxis_final.encode("utf-8")

                    st.download_button(
                        label="üì• Descargar sintaxis",
                        data=contenido_descarga,
                        file_name="sintaxis_busqueda.txt",
                        mime="text/plain"
                    )

        # --------------------- 4. PREPARAR RUTA ALMAC√âN V2 ---------------------
        def preparar_ruta():

            import streamlit as st
            import pandas as pd
            from io import BytesIO
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            st.title("üì¶ PREPARAR RUTA ALMAC√âN LA PAZ")

            col1, col2 = st.columns(2)
            file1 = col1.file_uploader("üìÇ Cargar Archivo 1", type="xlsx")
            file2 = col2.file_uploader("üìÇ Cargar Archivo 2", type="xlsx")

            with col1:
                file_name_input = st.text_input("üìù Nombre del archivo para CORREO", value="Archivo_Enviar_CORREO")
                file_name_correo = st.text_input("üìù Nombre archivo para RUTA", value="Archivo_Enviar_Ruta")
                sheet_name_input = st.text_input("üìÑ Nombre de la hoja", value="Ruta")
                
            with col2:
                
                recojo_title = st.text_input("‚úèÔ∏è T√≠tulo Recojo", value="RECOJO DE CAJAS B_SOL")
                devolucion_title = st.text_input("‚úèÔ∏è T√≠tulo Devoluci√≥n", value="DEVOLUCI√ìN DE ITEMS B_SOL")

            if file1 and file2:

                df1 = pd.read_excel(file1)
                df1.drop(['Estado de Solicitud', 'Servicio', 'Turno', 'Tipo de Recojo', 'Fecha de Registro'], 
                        axis=1, inplace=True, errors='ignore')

                df2 = pd.read_excel(file2)

                df1['SolicitudCode'] = df1['SolicitudCode'].astype(str).str.strip()
                df2['Nro Solicitud'] = df2['Nro Solicitud'].astype(str).str.strip()

                merged_df = pd.merge(
                    df1,
                    df2[['Nro Solicitud', 'Solicitante', 'Usuario', 'Items Oneil']],
                    left_on='SolicitudCode',
                    right_on='Nro Solicitud',
                    how='left'
                )

                if 'Centro de Costo' in merged_df.columns:
                    merged_df = merged_df.sort_values(by='Centro de Costo')

                df_recojo = merged_df[merged_df['TipoFile'] != 'FILE'].copy()
                df_devolucion = merged_df[merged_df['TipoFile'] == 'FILE'].copy()
                
                # üîç MOSTRAR PRIMERAS 5 FILAS
                st.subheader("üîç VISTA PREVIA PARA PREPARAR GUIA")
                st.dataframe(merged_df.head())

                st.subheader("üì¶ ‚Äì RECOJO")
                st.dataframe(df_recojo.head())

                st.subheader("üì¶ ‚Äì DEVOLUCI√ìN")
                st.dataframe(df_devolucion.head())

                for df in [df_recojo, df_devolucion, merged_df]:
                    df.drop(['Usuario', 'Nro Solicitud'], axis=1, inplace=True, errors='ignore')
                    df.dropna(how="all", inplace=True)
                    df.replace("", pd.NA, inplace=True)
                    df.dropna(how="all", inplace=True)

                # ======================================================
                # FUNCI√ìN BORDES
                # ======================================================
                thin_border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000")
                )

                def aplicar_bordes(ws, start_row, df):
                    nrows, ncols = df.shape
                    for row in range(start_row, start_row + nrows + 1):
                        for col in range(1, ncols + 1):
                            ws.cell(row=row, column=col).border = thin_border

                # ======================================================
                # FUNCI√ìN AUTOAJUSTE COLUMNAS
                # ======================================================
                def auto_ajustar_columnas(ws):
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = max_len + 3

                # ======================================================
                # ARCHIVO CORREO
                # ======================================================
                def generar_excel_principal(df_recojo, df_devolucion):
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        workbook = writer.book
                        ws = workbook.create_sheet(sheet_name_input)

                        if "Sheet" in workbook.sheetnames:
                            del workbook["Sheet"]

                        color_lila = "C27BA0"
                        color_rosa = "F4B1BA"
                        bold_font = Font(bold=True)

                        current_row = 1

                        # ------------------ RECOJO ------------------
                        last_col = get_column_letter(len(df_recojo.columns))
                        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
                        cell = ws[f"A{current_row}"]
                        cell.value = recojo_title
                        cell.font = Font(bold=True, size=14)
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill("solid", fgColor=color_lila)

                        startrow = current_row + 2
                        df_recojo.to_excel(writer, sheet_name=sheet_name_input, startrow=startrow, index=False)
                        header_row = startrow + 1

                        for col in range(1, len(df_recojo.columns) + 1):
                            h = ws.cell(row=header_row, column=col)
                            h.fill = PatternFill("solid", fgColor=color_lila)
                            h.font = bold_font

                        aplicar_bordes(ws, header_row, df_recojo)

                        current_row = header_row + len(df_recojo) + 3

                        # ------------------ DEVOLUCI√ìN ------------------
                        last_col = get_column_letter(len(df_devolucion.columns))
                        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
                        cell = ws[f"A{current_row}"]
                        cell.value = devolucion_title
                        cell.font = Font(bold=True, size=14)
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill("solid", fgColor=color_rosa)

                        startrow = current_row + 2
                        df_devolucion.to_excel(writer, sheet_name=sheet_name_input, startrow=startrow, index=False)
                        header_row = startrow + 1

                        for col in range(1, len(df_devolucion.columns) + 1):
                            h = ws.cell(row=header_row, column=col)
                            h.fill = PatternFill("solid", fgColor=color_rosa)
                            h.font = bold_font

                        aplicar_bordes(ws, header_row, df_devolucion)

                        # üî• Ajuste autom√°tico de columnas para todo el libro
                        auto_ajustar_columnas(ws)

                    output.seek(0)
                    return output

                # ======================================================
                # ARCHIVO RUTA
                # ======================================================
                def generar_excel_correo(df):
                    columnas = [
                        'SolicitudCode',
                        'Tipo de Solicitud',
                        'Cliente',
                        'WorkOrderCode',
                        'Cantidad',
                        'TipoFile',
                        'Centro de Costo',
                        'Solicitante',
                        'Items Oneil'
                    ]

                    df_simple = df[columnas].copy()

                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_simple.to_excel(writer, index=False, sheet_name="RUTA")
                        ws = writer.book["RUTA"]

                        aplicar_bordes(ws, 1, df_simple)
                        auto_ajustar_columnas(ws)   # üî• Autoajuste aqu√≠ tambi√©n

                    output.seek(0)
                    return output

                # ======================================================
                # DESCARGAS
                # ======================================================
                archivo_principal = generar_excel_principal(df_recojo, df_devolucion)
                archivo_correo = generar_excel_correo(merged_df)

                st.download_button(
                    "üì• Descargar Archivo para CORREO",
                    archivo_principal,
                    file_name=file_name_input + ".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.download_button(
                    "üìß Descargar Archivo para RUTA",
                    archivo_correo,
                    file_name=file_name_correo + ".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


        # --------------------- 5. PREPARAR GUIA PARA RUTA V2 --------------------- 
        def preparar_guia_ruta():
              
            import streamlit as st
            import pandas as pd
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            import datetime

            st.title("üì¶ GENERADOR DE GU√çA - FORMATO OFICIAL POLYSISTEMAS")

            # Cargar archivos
            col1, col2 = st.columns(2)
            with col1:
                logo_file = st.file_uploader("üì∏ Subir Logo (PNG o JPG)", type=["png", "jpg", "jpeg"])
            with col2:
                uploaded_file = st.file_uploader("Cargar archivo Excel para generar la Ruta", type="xlsx")

            if uploaded_file:
                df = pd.read_excel(uploaded_file)
                df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("-", "_")

                required_columns = ['Solicitante', 'Centro_de_Costo', 'WorkOrderCode', 'Cantidad', 'Items_Oneil', 'TipoFile']
                if not all(col in df.columns for col in required_columns):
                    st.error(f"‚ùå Faltan columnas requeridas: {', '.join(required_columns)}")
                    st.stop()

                # ‚úî Convertir None a "CINTILLOS"
                df["Items_Oneil"] = df["Items_Oneil"].fillna("CINTILLOS")

                # Selectores
                col1, col2, col3, col4 = st.columns(4)

                with col1:
                    personal = st.selectbox("üë§ Personal de Polysistemas:",
                                            ["JAIME QUISPE", "CARLOS ORTIZ", "MARCO HUAYLLUCO", "ALFREDO RIVEROS"])

                with col2:
                    cliente = st.selectbox("üè¢ Cliente:",
                                        ["BANCO SOL", "BNB", "BANCO FIE"])

                with col3:
                    fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y")

                    tipo_seleccionado = st.multiselect(
                        "Selecciona uno o m√°s TipoFile a procesar:",
                        ["CAJA", "Caja", "Cintillos", "FILE"],
                        default=["CAJA"]
                    )

                with col4:
                    regional = st.selectbox("üåç REGIONAL:", ["LA PAZ", "EL ALTO", "ORURO", "POTOSI", "SUCRE", "VIACHA"])

                # Filtrar TipoFile
                df = df[df["TipoFile"].isin(tipo_seleccionado)]

                st.info(f"""
                **üë§ Personal:** {personal}  
                **üè¢ Cliente:** {cliente}  
                **üìÖ Fecha:** {fecha_actual}  
                **üéØ TipoFile Seleccionado:** {', '.join(tipo_seleccionado)}  
                """)

                st.subheader("üîç Vista previa de datos filtrados")
                st.write(df.head(5))

                # FUNCI√ìN PARA RANGO SOLO PARA "Caja"
                def obtener_rango_codigos(cadena, tipo_file):
                    # Solo aplicar rango si es exactamente "Caja"
                    if tipo_file != "Caja":
                        if cadena in [None, "CINTILLOS"]:
                            return "CINTILLOS"
                        return str(cadena)

                    # Si es Caja ‚Üí aplicar rango
                    if cadena in [None, "CINTILLOS"]:
                        return "CINTILLOS"

                    partes = [p.strip() for p in str(cadena).replace(";", ",").split(",") if p.strip().isdigit()]

                    if not partes:
                        return "CINTILLOS"

                    partes = sorted(int(p) for p in partes)
                    return f"{partes[0]} - {partes[-1]}"

                # Crear Excel
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Gu√≠a"
                ws.sheet_view.showGridLines = False

                # Estilos
                bold = Font(bold=True)
                white_bold = Font(color="FFFFFF", bold=True)
                center = Alignment(horizontal="center", vertical="center")
                left = Alignment(horizontal="left", vertical="center")

                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                rojo = PatternFill("solid", fgColor="933C47")
                gris = PatternFill("solid", fgColor="F2F2F2")

                def border_full_row(row, cols=7):
                    for c in range(1, cols + 1):
                        ws.cell(row, c).border = border

                # Logo
                if logo_file:
                    img = XLImage(logo_file)
                    img.width = 190
                    img.height = 75
                    ws.add_image(img, "A1")

                # ENCABEZADOS
                ws.merge_cells("C1:E1")
                ws["C1"].value = "FORMATO"
                ws["C1"].font = bold
                ws["C1"].alignment = center

                for col in range(3, 6):
                    ws.cell(row=1, column=col).border = border

                ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=5)
                ws["C2"].value = "GUIA DE RECEPCI√ìN DE MENSAJER√çA"
                ws["C2"].font = bold
                ws["C2"].alignment = center

                for col in range(3, 6):
                    ws.cell(row=2, column=col).border = border
                    ws.cell(row=3, column=col).border = border

                encabezado_pairs = {
                    
                    "F1": "C√≥digo:", "G1": "BOL-2.18-FOR-01",
                    "F2": "Edici√≥n:", "G2": "02",
                    "F3": "Fecha Edici√≥n:", "G3": "26/11/2025"
                }

                for cell, value in encabezado_pairs.items():
                    ws[cell].value = value
                    ws[cell].alignment = center
                    ws[cell].border = border
                    if "F" in cell:
                        ws[cell].font = bold

                # PERSONAL / FECHA
                ws.merge_cells("A5:C5")
                ws["A5"].value = "PERSONAL DE POLYSISTEMAS"
                ws["A5"].font = white_bold
                ws["A5"].fill = rojo

                ws.merge_cells("D5:E5")
                ws["D5"].value = personal
                ws["D5"].alignment = center
                ws["D5"].font = bold

                ws["F5"].value = "FECHA:"
                ws["F5"].font = white_bold
                ws["F5"].fill = rojo

                ws["G5"].value = fecha_actual
                ws["G5"].alignment = center
                ws["G5"].font = bold

                border_full_row(5)

                # CLIENTE
                ws.merge_cells("A6:C6")
                ws["A6"].value = "CLIENTE:"
                ws["A6"].font = white_bold
                ws["A6"].fill = rojo

                ws.merge_cells("D6:E6")
                ws["D6"].value = cliente
                ws["D6"].alignment = center
                ws["D6"].font = bold

                ws["F6"].value = "CANTIDAD TOTAL:"
                ws["F6"].font = white_bold
                ws["F6"].fill = rojo

                total_cantidad = df["Cantidad"].sum()
                ws["G6"].value = total_cantidad
                ws["G6"].alignment = center
                ws["G6"].font = bold

                border_full_row(6)

                # REGIONAL
                ws.merge_cells("A8:B8")
                ws["A8"].value = "REGIONAL:"
                ws["A8"].font = bold
                ws["A8"].fill = gris

                ws.merge_cells("C8:E8")
                ws["C8"].value = regional
                ws["C8"].alignment = center
                ws["C8"].font = bold

                border_full_row(8)

                # TABLAS POR SOLICITANTE
                start_row = 9
                solicitantes = df["Solicitante"].unique()

                for solicitante in solicitantes:
                    df_s = df[df["Solicitante"] == solicitante]
                    agencia = df_s["Centro_de_Costo"].iloc[0]

                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
                    ws.cell(start_row, 1).value = "SOLICITANTE:"
                    ws.cell(start_row, 1).fill = gris
                    ws.cell(start_row, 1).font = bold

                    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=5)
                    ws.cell(start_row, 3).value = solicitante
                    ws.cell(start_row, 3).font = bold
                    ws.cell(start_row, 3).alignment = center

                    ws.cell(start_row, 6).value = "AGENCIA:"
                    ws.cell(start_row, 6).fill = gris
                    ws.cell(start_row, 6).font = bold

                    ws.cell(start_row, 7).value = agencia
                    ws.cell(start_row, 7).font = bold
                    ws.cell(start_row, 7).alignment = center

                    border_full_row(start_row)
                    start_row += 1

                    headers = ["NRO", "NRO. WO", "CANTIDAD", "CONTAINER CODE", "OBSERVACIONES", "FIRMA ENTREGA", "FIRMA RECEPCI√ìN"]
                    for col, h in enumerate(headers, 1):
                        ws.cell(start_row, col).value = h
                        ws.cell(start_row, col).font = white_bold
                        ws.cell(start_row, col).fill = rojo
                        ws.cell(start_row, col).alignment = center

                    border_full_row(start_row)
                    start_row += 1

                    nro = 1
                    total_solic = 0

                    for r in df_s.itertuples(index=False):
                        ws.cell(start_row, 1, nro).alignment = center
                        ws.cell(start_row, 2, r.WorkOrderCode).alignment = center
                        ws.cell(start_row, 3, r.Cantidad).alignment = center

                        # ‚úî SOLO aplica rango cuando TipoFile == "Caja"
                        codigo_rango = obtener_rango_codigos(r.Items_Oneil, r.TipoFile)
                        ws.cell(start_row, 4, codigo_rango).alignment = center

                        border_full_row(start_row)

                        total_solic += r.Cantidad
                        nro += 1
                        start_row += 1

                    # Totales
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
                    ws.cell(start_row, 1).value = "TOTAL"
                    ws.cell(start_row, 1).font = bold
                    ws.cell(start_row, 1).alignment = center

                    ws.cell(start_row, 3).value = total_solic
                    ws.cell(start_row, 3).alignment = center

                    ws.cell(start_row, 6).value = "ENTREGU√â CONFORME"
                    ws.cell(start_row, 6).fill = gris
                    ws.cell(start_row, 6).font = bold
                    ws.cell(start_row, 6).alignment = center

                    ws.cell(start_row, 7).value = "RECIB√ç CONFORME"
                    ws.cell(start_row, 7).fill = gris
                    ws.cell(start_row, 7).font = bold
                    ws.cell(start_row, 7).alignment = center

                    border_full_row(start_row)
                    start_row += 2

                # Ajuste columnas
                widths = [10, 20, 12, 40, 30, 22, 22]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="üì• Descargar Gu√≠a en Formato Oficial",
                    data=output,
                    file_name="Guia_Polysistemas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

          
        # --------------------- 6. ELIMINAR ARCHIVOS ---------------------
        def eliminar_archivos():
            import streamlit as st
            import os
            import glob
            import pandas as pd

            #st.set_page_config(page_title="Eliminar Excel con vista previa", layout="centered")
            st.title("üóëÔ∏è Eliminar archivos Excel con vista previa desplegable")

            folder_input = st.text_input("Introduce la ruta de la carpeta con archivos Excel:")
            if st.button("üîÑ Cargar archivos"):
                st.session_state.folder = folder_input
                st.session_state.selected_files = []

            # Inicializar estados
            if 'folder' not in st.session_state:
                st.session_state.folder = ""
            if 'selected_files' not in st.session_state:
                st.session_state.selected_files = []

            folder = st.session_state.folder

            def toggle_selection(f_name):
                sel = st.session_state.selected_files
                if f_name in sel:
                    sel.remove(f_name)
                else:
                    sel.append(f_name)

            if folder:
                if os.path.isdir(folder):
                    files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
                    if files:
                        st.subheader("üìÇ Selecciona archivos:")
                        for path in files:
                            name = os.path.basename(path)
                            selected = name in st.session_state.selected_files
                            col1, col2 = st.columns([4,1])
                            with col1:
                                st.write(name)
                            with col2:
                                if st.button("‚úÖ" if selected else "‚òê", key=name):
                                    toggle_selection(name)
                                    st.rerun()

                        if st.session_state.selected_files:
                            st.subheader("üëÅÔ∏è Previsualizaciones desplegables")
                            for name in st.session_state.selected_files:
                                file_path = os.path.join(folder, name)
                                with st.expander(f"Ver preview: {name}", expanded=False):
                                    try:
                                        df = pd.read_excel(file_path, nrows=10)
                                        st.dataframe(df)
                                    except Exception as e:
                                        st.error(f"Error leyendo {name}: {e}")

                            if st.checkbox("‚ö†Ô∏è Confirmo que deseo eliminar los archivos seleccionados"):
                                if st.button("üóëÔ∏è Eliminar archivos"):
                                    cnt = 0
                                    for name in list(st.session_state.selected_files):
                                        p = os.path.join(folder, name)
                                        if os.path.exists(p):
                                            os.remove(p)
                                            cnt += 1
                                    st.success(f"Se eliminaron {cnt} archivos correctamente.")
                                    st.session_state.selected_files = []
                                    st.session_state.folder = ""
                                    st.rerun()
                        else:
                            st.info("Selecciona archivos pulsando en los botones de la derecha.")
                    else:
                        st.warning("No se encontraron archivos Excel en la carpeta.")
                else:
                    st.error("La ruta no es v√°lida o no corresponde a una carpeta.")
            else:
                st.info("Introduce la ruta y pulsa 'Cargar archivos' para empezar.")

                                
        # --------------------- EJECUTAR SEG√öN MEN√ö ---------------------
        if menu == "üè∑Ô∏è Filtrar Solicitudes":
            filtrar_solicitudes()
        elif menu == "üß© Combinar Archivos":
            combinar_archivos()
        elif menu == "üîé Generar Sintaxis de B√∫squeda":
            generar_sintaxis()
        elif menu == "üì¶ Preparar Ruta Almac√©n La Paz":
            preparar_ruta()
        elif menu == "üì¶ Preparar Guias de Ruta":
            preparar_guia_ruta()
        elif menu == "üóëÔ∏è Eliminar Archivos":
            eliminar_archivos()


elif proyecto == "CHECK FILEWEB AND LASERFICHE":
    import streamlit as st
    import pandas as pd
    import os
    from io import BytesIO

    st.subheader("üîç CHECK FILEWEB AND LASERFICHE")

    # --------------------- SIDEBAR MENU ---------------------
    st.sidebar.title("Men√∫ de Opciones")
    menu = st.sidebar.radio("Ir a:", [
        "üîé Generar Sintaxis Solicitudes",
        "üè∑Ô∏è Reemplazar archivo",
        "üß© Comparaci√≥n de Solicitudes FileWeb vs LaserFiche",
        "üîé Generar Sintaxis LaserFiche del filtrado",
        "üîé Sintaxis ONEIL del filtrado",
        "üóëÔ∏è Eliminar Archivos Generados"

    ])
        
    # --------------------- 1. GENERADOR DE SINTAXIS ---------------------
    def Generar_Sintaxis_Solicitudes():

        st.title("üîé GENERADOR DE SINTAXIS DE B√öSQUEDA LASER FICHE")

        if "sintaxi_final" not in st.session_state:
            st.session_state.sintaxis_final = ""
            st.session_state.sistaxis_preview = ""
            st.session_state.total_registros = 0

        col1, col2 = st.columns((2))
        with col1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")


            # Generar la sintaxis de b√∫squeda
            if st.button("Generar Sintaxis"):

                if uploaded_file is not None:
                    try:
                        # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                        df = pd.read_excel(uploaded_file)

                        # Verificar si la columna existe
                        if 'SolicitudCodeFileWeb' not in df.columns:
                            st.error("La columna 'SolicitudCodeFileWeb' no se encuentra en el archivo.")
                        else:
                            # Eliminar las dem√°s columnas y eliminar nulos
                            solicitud_codes = df['SolicitudCodeFileWeb'].dropna().astype(str).tolist()

                            # Crear la sintaxis
                            #solicitud_sintaxis = " | ".join(
                                #[f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                
                            #)
                            # üîπ Sintaxis completa
                            sintaxis_list = [
                                f'{{[Orden de Trabajo]:[Nro Solicitud]="{code.strip()}"}}'
                                for code in solicitud_codes if code.strip()
                            ]

                            st.session_state.sintaxis_final = " | \n".join(sintaxis_list)

                            # üîπ Preview: primeros 20
                            st.session_state.sintaxis_preview = " | \n".join(sintaxis_list[:20])
                            
                            st.session_state.total_registros = len(sintaxis_list)

                            st.success("Sintaxis generada correctamente")


                    except Exception as e:
                        st.error(f"Error al leer el archivo: {e}")

            # Pie de p√°gina
            st.markdown("---")
            st.caption("‚ÑπÔ∏è PARA DESCARGAR DE LASER FICHE DEL CAMPO (Orden de Trabajo) AGREGAR [Nro de Solicitus y Nro de WO] ")

        with col2:
            # Mostrar la sintaxis generada
            if st.session_state.sintaxis_final:
                # üîπ Mostrar total
                st.info(f"üìä Total de registros generados: **{st.session_state.total_registros}**")

                st.success("Sintaxis de b√∫squeda generada:")
                st.code(st.session_state.sintaxis_preview)

                # Boton de descarga 
                st.download_button(
                    label="‚¨áÔ∏è Descargar Sintaxis",
                    data=st.session_state.sintaxis_final,
                    file_name="Sintaxis_LaserFiche.txt",
                    mime="text/plain"
                )


    # --------------------- 2. Reemplazar Archivos ---------------------

    def Reemplazar_Archivo():
        import streamlit as st
        import pandas as pd
        import os
        import io

        #st.set_page_config(page_title="Reemplazo de Archivos Excel", page_icon="üìä", layout="wide")
        st.title("Reemplazo de Archivos Excel üóÇÔ∏è")

        # Funci√≥n para cargar los archivos Excel
        def cargar_archivo(nombre):
            """Cargar un archivo Excel en un DataFrame y devolver el nombre del archivo."""
            archivo = st.file_uploader(f"Sube el archivo {nombre}", type=["xlsx", "xls"])
            if archivo is not None:
                # Obtener el nombre del archivo
                nombre_archivo = archivo.name
                # Cargar el archivo Excel en un DataFrame de Pandas
                df = pd.read_excel(archivo, engine='openpyxl')
                return df, nombre_archivo
            return None, None

        # Funci√≥n para reemplazar los datos
        def reemplazar_datos(df1, df2):
            """Reemplazar solo los datos del archivo 1 con los datos del archivo 2, manteniendo el encabezado."""
            # Guardar el encabezado del archivo 1
            encabezado_df1 = df1.columns
            
            # Reemplazar los datos, manteniendo el encabezado del archivo 1
            df2.columns = encabezado_df1  # Aseguramos que los encabezados de df2 coincidan con df1
            df2 = df2[encabezado_df1]     # Solo tomamos las columnas de df2 que est√°n en df1
            return df2

        # Funci√≥n para guardar el archivo reemplazado en la ruta proporcionada
        def guardar_archivo(df, ruta, nombre_archivo):
            """Guardar el archivo reemplazado en la ruta especificada por el usuario."""
            if not os.path.exists(ruta):
                st.error("üö® ¬°La ruta no existe! Por favor, ingresa una ruta v√°lida.")
                return None
            
            # Guardar el archivo como Excel en la ruta proporcionada
            ruta_completa = os.path.join(ruta, nombre_archivo)
            df.to_excel(ruta_completa, index=False, engine='openpyxl')
            
            return ruta_completa

        # Subir los dos archivos Excel
        col1, col2 = st.columns(2)
        with col1:
            archivo_1_df, nombre_archivo_1 = cargar_archivo("1")
        with col2:
            archivo_2_df, _ = cargar_archivo("2")


        # Si los archivos han sido subidos
        if archivo_1_df is not None and archivo_2_df is not None:
            # Reemplazar los datos del archivo 1 con los del archivo 2, manteniendo el encabezado de archivo 1
            archivo_reemplazado = reemplazar_datos(archivo_1_df, archivo_2_df)
            
            # Mostrar los primeros 5 registros del archivo reemplazado
            st.subheader("Archivo Reemplazado (Vista Previa) üëÄ")
            st.write(archivo_reemplazado.head())
            
            # Crear archivo Excel en memoria
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                archivo_reemplazado.to_excel(writer, index=False)
            buffer.seek(0)

            # Bot√≥n de descarga
            st.download_button(
                label="üì• Descargar archivo Excel",
                data=buffer,
                file_name=nombre_archivo_1,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    # --------------------- 3. Comparaci√≥n de Solicitudes FileWeb vs LaserFiche ---------------------
    def Comparar_Solicitudes_FW_y_LF():
        import pandas as pd
        import streamlit as st
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
        from collections import defaultdict
        import os

        st.title("üìã Comparaci√≥n de Solicitudes FileWeb vs LaserFiche")

        def procesar_archivos(file1, file2):
            df1 = pd.read_excel(file1)
            df2 = pd.read_excel(file2)

            # NORMALIZAR COLUMNAS (CLAVE)
            df1.columns = (
                df1.columns
                .str.replace('\n', ' ', regex=True)
                .str.replace('\xa0', ' ', regex=True)
                .str.strip()
            )

            df2.columns = (
                df2.columns
                .str.replace('\n', ' ', regex=True)
                .str.replace('\xa0', ' ', regex=True)
                .str.strip()
            )

            #st.write("Columnas LaserFiche detectadas:", list(df2.columns))

            # Asegurar que las columnas relevantes sean texto
            columnas_texto = [
                'SolicitudCodeFileWeb', 'Tipo de Solicitud FileWeb', 'Estado de Solicitud FileWeb',
                'WorkOrderCodeFileWe', 'CantidadFileWeb', 'TipoFileFileWeb', 'Centro de Costo FileWeb',
                'Fecha de Registro FileWeb', 'NombreLaserFiche', 'Cliente LaserFiche',
                'Estado de OT LaserFiche', 'Nro Solicitud laserFiche', 'Direccion'
            ]
            for df in [df1, df2]:
                for col in columnas_texto:
                    if col in df.columns:
                        df[col] = df[col].astype(str).fillna("")

            merged_df = pd.merge(
                df1[['SolicitudCodeFileWeb', 'Tipo de Solicitud FileWeb', 'Estado de Solicitud FileWeb',
                    'WorkOrderCodeFileWe', 'CantidadFileWeb', 'TipoFileFileWeb', 'Centro de Costo FileWeb',
                    'Fecha de Registro FileWeb']],
                df2[['NombreLaserFiche', 'Cliente LaserFiche', 'Estado de OT LaserFiche', 'Nro Solicitud laserFiche', 'Direccion']],
                left_on='SolicitudCodeFileWeb',
                right_on='Nro Solicitud laserFiche',
                how='inner'
            )

            column_order = [
                'SolicitudCodeFileWeb',
                'Nro Solicitud laserFiche',
                'Tipo de Solicitud FileWeb',
                'NombreLaserFiche',
                'Estado de Solicitud FileWeb',
                'Estado de OT LaserFiche',
                'WorkOrderCodeFileWe',
                'Fecha de Registro FileWeb',
                'CantidadFileWeb',
                'TipoFileFileWeb',
                'Centro de Costo FileWeb',
                'Cliente LaserFiche',
                'Direccion'
            ]

            return merged_df[column_order]

        # Subida de archivos
        col1, col2 = st.columns(2)
        file1 = col1.file_uploader("üìÇ Sube archivo FileWeb", type=["xlsx"])
        file2 = col2.file_uploader("üìÇ Sube archivo LaserFiche", type=["xlsx"])

        if file1 and file2:
            try:
                resultado = procesar_archivos(file1, file2)

                # Comparar estados de forma segura
                resultado['Rojo'] = resultado.apply(
                    lambda row: str(row['Estado de Solicitud FileWeb']).strip() != str(row['Estado de OT LaserFiche']).strip(),
                    axis=1
                )

                # Opci√≥n de mostrar todo o solo diferentes
                opcion = st.radio("¬øQu√© archivo deseas ver y descargar?", ("Todo", "Solo Rojo"))

                if opcion == "Solo Rojo":
                    resultado = resultado[resultado['Rojo'] == True]

                # Filtros
                estados_ot = sorted(resultado['Estado de OT LaserFiche'].dropna().unique())
                estado_ot = st.selectbox("üìå Filtrar por Estado de OT LaserFiche:", ["(Todos)"] + estados_ot)

                estados_fileweb = sorted(resultado['Estado de Solicitud FileWeb'].dropna().unique())
                estado_fileweb = st.selectbox("üìç Filtrar por Estado de Solicitud FileWeb:", ["(Todos)"] + estados_fileweb)

                prefijo = st.text_input("üî§ Buscar por prefijo de NombreLaserFiche (ej: OTEV, OTPM):").strip().upper()

                # Aplicar filtros
                if estado_ot != "(Todos)":
                    resultado = resultado[resultado['Estado de OT LaserFiche'] == estado_ot]

                if estado_fileweb != "(Todos)":
                    resultado = resultado[resultado['Estado de Solicitud FileWeb'] == estado_fileweb]

                if prefijo:
                    resultado = resultado[resultado['NombreLaserFiche'].str.upper().str.startswith(prefijo)]

                st.success(f"‚úÖ Se encontraron {len(resultado)} registros filtrados")
                st.dataframe(resultado.drop(columns=['Rojo']), height=600, use_container_width=True)

                # Exportar a Excel con formato
                resultado_exportar = resultado.drop(columns=['Rojo'])
                excel_buffer = BytesIO()
                resultado_exportar.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                wb = load_workbook(excel_buffer)
                ws = wb.active
                ws.auto_filter.ref = ws.dimensions

                # Estilos
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center_align = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border

                # Mapeo columnas
                header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1)}
                col_estado1 = header_map.get("Estado de Solicitud FileWeb")
                col_estado2 = header_map.get("Estado de OT LaserFiche")
                col_workorder = header_map.get("WorkOrderCodeFileWe")

                verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                rojo = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
                intercalados = [PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
                                PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")]

                # Colorear diferencias
                for row in range(2, ws.max_row + 1):
                    val1 = ws.cell(row=row, column=col_estado1).value
                    val2 = ws.cell(row=row, column=col_estado2).value
                    if val1 and val2:
                        if str(val1).strip() != str(val2).strip():
                            ws.cell(row=row, column=col_estado1).fill = rojo
                            ws.cell(row=row, column=col_estado2).fill = rojo
                        else:
                            ws.cell(row=row, column=col_estado1).fill = verde
                            ws.cell(row=row, column=col_estado2).fill = verde

                # Resaltar duplicados
                if col_workorder:
                    duplicados = defaultdict(list)
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=col_workorder).value
                        if val:
                            duplicados[str(val).strip()].append(row)
                    grupos = [v for v in duplicados.values() if len(v) > 1]
                    for i, filas in enumerate(grupos):
                        fill = intercalados[i % 2]
                        for row in filas:
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = fill

                # Ajustar ancho y aplicar bordes
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = thin_border
                    ws.column_dimensions[col_letter].width = max_length + 2

                final_buffer = BytesIO()
                wb.save(final_buffer)
                final_buffer.seek(0)

                # Guardar localmente si se indica
                ruta_personalizada = st.text_input("üìÅ Ingrese una ruta de servidor para guardar el archivo (opcional):")
                if ruta_personalizada:
                    try:
                        output_path = os.path.join(ruta_personalizada, "comparacion_de_solicitudes.xlsx")
                        with open(output_path, "wb") as f:
                            f.write(final_buffer.getbuffer())
                        st.success(f"‚úÖ Archivo guardado en: {output_path}")
                    except Exception as e:
                        st.error(f"‚ùå No se pudo guardar el archivo: {str(e)}")

                st.download_button(
                    label="üíæ Descargar archivo con formato",
                    data=final_buffer,
                    file_name="comparacion_de_solicitudes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"‚ùå Error: {str(e)}")

        st.markdown("---")
        st.caption("üîé Este archivo contiene filtros, bordes, colores para diferencias y filas duplicadas.")


    def Generar_Sintaxis_del_filtrado():

        st.title("üîé GENERADOR DE SINTAXIS PARA LASER FICHE DEL ARCHIVO FILTRADO")

        colum1, colum2 = st.columns((2))
        with colum1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

            # Variable para almacenar la sintaxis generada
            sintaxis_final = ""

            # Generar la sintaxis de b√∫squeda
            if st.button("Generar Sintaxis"):

                if uploaded_file is not None:
                    try:
                        # Leer el archivo Excel y obtener solo la columna 'SolicitudCode'
                        df = pd.read_excel(uploaded_file)

                        # Verificar si la columna existe
                        if 'SolicitudCodeFileWeb' not in df.columns:
                            st.error("La columna 'SolicitudCodeFileWeb' no se encuentra en el archivo.")
                        else:
                            # Eliminar las dem√°s columnas y eliminar nulos
                            solicitud_codes = df['SolicitudCodeFileWeb'].dropna().astype(str).tolist()

                            # Crear la sintaxis
                            solicitud_sintaxis = " | ".join(
                                [f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{code.strip()}"}}' for code in solicitud_codes if code.strip()]
                                
                            )

                            sintaxis_final = f"{solicitud_sintaxis}"
                    except Exception as e:
                        st.error(f"Error al leer el archivo: {e}")

        with colum2:
            # Mostrar la sintaxis generada
            if sintaxis_final:
                st.success("Sintaxis de b√∫squeda generada:")
                st.code(sintaxis_final)
        
        # Pie de p√°gina
        st.markdown("---")

    def Generar_Sintaxis_ONEIL_del_filtrado():
        import streamlit as st
        import pandas as pd

        st.title("üîé GENERADOR DE SINTAXIS PARA ONEIL DEL ARCHIVO FILTRADO")
        colums1, colums2 = st.columns((2))

        with colums1:
            # Cargar archivo Excel
            st.markdown("<h4>Selecciona el archivo Excel con la columna 'SolicitudCodeFileWeb'</h4>", unsafe_allow_html=True)
            uploaded_files = st.file_uploader("Selecciona un archivo .xlsx", type="xlsx")

            # Variable para almacenar la sintaxis generada
            sintaxis_final = ""

            if st.button("Generar Sintaxis"):
                if uploaded_files is not None:
                    # Leer el archivo Excel con pandas
                    try:
                        df = pd.read_excel(uploaded_files)

                        # Verificar si la columna existe
                        if "SolicitudCodeFileWeb" in df.columns:
                            # Eliminar valores vac√≠os y construir la sintaxis
                            solicitud_numbers = df["SolicitudCodeFileWeb"].dropna().astype(str)

                            solicitud_sintaxis = " OR ".join([
                                f'\n Workorder^PONumber = "{solicitud.strip()}"'
                                for solicitud in solicitud_numbers
                            ])
                            sintaxis_final = f"{solicitud_sintaxis}"
                        else:
                            st.error("‚ùå La columna 'SolicitudCodeFileWeb' no se encontr√≥ en el archivo.")
                    except Exception as e:
                        st.error(f"‚ùå Error al leer el archivo: {e}")

        with colums2:
            if sintaxis_final:
                st.success("Sintaxis de b√∫squeda generada:")
                st.code(sintaxis_final)

    
   
    # ------************************************************************************------

    def eliminar_archivos_generados():
            import streamlit as st
            import os
            import glob
            import pandas as pd

            #st.set_page_config(page_title="Eliminar Excel con vista previa", layout="centered")
            st.title("üóëÔ∏è Eliminar archivos Excel con vista previa desplegable")

            folder_input = st.text_input("Introduce la ruta de la carpeta con archivos Excel:", "C:\\Users\\juan.ramos\\Desktop\\FILTRADOR_ALMACEN\\CHECK_FILEWEB_AND_LASER_FICHE")
            if st.button("üîÑ Cargar archivos"):
                st.session_state.folder = folder_input
                st.session_state.selected_files = []

            # Inicializar estados
            if 'folder' not in st.session_state:
                st.session_state.folder = ""
            if 'selected_files' not in st.session_state:
                st.session_state.selected_files = []

            folder = st.session_state.folder

            def toggle_selection(f_name):
                sel = st.session_state.selected_files
                if f_name in sel:
                    sel.remove(f_name)
                else:
                    sel.append(f_name)

            if folder:
                if os.path.isdir(folder):
                    files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
                    if files:
                        st.subheader("üìÇ Selecciona archivos:")
                        for path in files:
                            name = os.path.basename(path)
                            selected = name in st.session_state.selected_files
                            col1, col2 = st.columns([4,1])
                            with col1:
                                st.write(name)
                            with col2:
                                if st.button("‚úÖ" if selected else "‚òê", key=name):
                                    toggle_selection(name)
                                    st.rerun()

                        if st.session_state.selected_files:
                            st.subheader("üëÅÔ∏è Previsualizaciones desplegables")
                            for name in st.session_state.selected_files:
                                file_path = os.path.join(folder, name)
                                with st.expander(f"Ver preview: {name}", expanded=False):
                                    try:
                                        df = pd.read_excel(file_path, nrows=10)
                                        st.dataframe(df)
                                    except Exception as e:
                                        st.error(f"Error leyendo {name}: {e}")

                            if st.checkbox("‚ö†Ô∏è Confirmo que deseo eliminar los archivos seleccionados"):
                                if st.button("üóëÔ∏è Eliminar archivos"):
                                    cnt = 0
                                    for name in list(st.session_state.selected_files):
                                        p = os.path.join(folder, name)
                                        if os.path.exists(p):
                                            os.remove(p)
                                            cnt += 1
                                    st.success(f"Se eliminaron {cnt} archivos correctamente.")
                                    st.session_state.selected_files = []
                                    st.session_state.folder = ""
                                    st.rerun()
                        else:
                            st.info("Selecciona archivos pulsando en los botones de la derecha.")
                    else:
                        st.warning("No se encontraron archivos Excel en la carpeta.")
                else:
                    st.error("La ruta no es v√°lida o no corresponde a una carpeta.")
            else:
                st.info("Introduce la ruta y pulsa 'Cargar archivos' para empezar.")


    # --------------------- EJECUTAR SEG√öN MEN√ö ---------------------
    if menu == "üîé Generar Sintaxis Solicitudes":
        Generar_Sintaxis_Solicitudes()
    elif menu == "üß© Comparaci√≥n de Solicitudes FileWeb vs LaserFiche":
        Comparar_Solicitudes_FW_y_LF()
    elif menu == "üè∑Ô∏è Reemplazar archivo":
        Reemplazar_Archivo()
    elif menu == "üîé Generar Sintaxis LaserFiche del filtrado":
        Generar_Sintaxis_del_filtrado()
    elif menu == "üîé Sintaxis ONEIL del filtrado":
        Generar_Sintaxis_ONEIL_del_filtrado()
    elif menu == "üóëÔ∏è Eliminar Archivos Generados":
        eliminar_archivos_generados()

elif proyecto == "SINTAXIS LASERFICHE-ONEIL":
    import streamlit as st
    from datetime import datetime
    
    st.subheader("üß† SINTAXIS LASERFICHE-ONEIL")
    
    # --------------------- SIDEBAR MENU ---------------------
    st.sidebar.title("Men√∫ de Opciones")
    menu = st.sidebar.radio("Ir a:", [
        "üß± OTPM - MATERIALES",
        "üóÉÔ∏è OTEX - EXTRACIONES",
        "üì§ OTEV - ENVIO",
        "üì• OTRE - RECOJO",
        "‚úâÔ∏è OTRE - POR_WORKORDERS",
        "üßæ OTRE - POR_SOLICITUDES",
        "üìù OTRE - SOLICITUDES",
        "üß† COD - BASE_DE_DATO",
        "üé® CONTENIDO_PALLETS",
        "üîñ CODIGOS DE 6 DIGITOS"
        
    ])
    def OTPM_MATERIALES():
        
        opciones2 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "PENDIENTE":
            #import streamlit as st
            from datetime import datetime
            
            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparaci√≥n de Materiales\\03. Finalizada\\01. Pendiente"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()
    
        elif seleccion2 == "EN PROCESO":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparaci√≥n de Materiales\\02. En Proceso"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()

        elif seleccion2 == "FINALIZADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparaci√≥n de Materiales\\03. Finalizada"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()
        
        elif seleccion2 == "ANULADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\01. OTPM - Preparaci√≥n de Materiales\\04. Anulada"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                        
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()

    def OTEX_EXTRACIONES():

        # Crear una lista de opciones para el menu 
        opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

        # Crear el menu desplegable con st.selectbox()
        seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

        if seleccion3 == "PENDIENTE":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracci√≥n\\01. Pendiente"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()


        elif seleccion3 == "EN PROCESO":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracci√≥n\\02. En proceso"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()

        elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracci√≥n\\03. Finalizada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

        elif seleccion3 == "ANULADA":
            #import streamlit as st
            from datetime import datetime

            def main():
                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Sintaxis predeterminada
                sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\06. OTEX - Extracci√≥n\\04. Anulada"}'

                # Secci√≥n para seleccionar una fecha usando un calendario
                st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                    # Formatear la fecha seleccionada
                    fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                    sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                    # Mostrar la sintaxis generada
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

            if __name__ == "__main__":
                main()

    def OTEV_ENVIO():

        # Crear una lista de opciones para el menu 
        opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "INTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Interno\\01. Pendiente"}'

                    # Secci√≥n para seleccionar una fecha usando un calendariO
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()


            elif seleccion2 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Interno\\02. En proceso"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion2 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Interno\\03. Finalizada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                        main()

            elif seleccion2 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Interno\\04. Anulada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

        if seleccion2 == "EXTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Externo\\01. Pendiente"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Externo\\02. En proceso"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Externo\\03. Finalizada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTEV - Env√≠o\\01. Solicitante Externo\\04. Anulada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

    def OTRE_RECOJO():

        # Crear una lista de opciones para el menu 
        opciones2 = ["SELECCIONA UNA OPCION", "INTERNO", "EXTERNO"]

        # Crear el menu desplegable con st.selectbox()
        seleccion2 = st.selectbox("Selecciona una opcion del menu: ", opciones2)

        if seleccion2 == "INTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\01. Pendiente"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.header("Seleccione una fecha")
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\02. En proceso"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\03. Finalizada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\01. Solicitante Interno\\04. Anulada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

        if seleccion2 == "EXTERNO":
            # Crear una lista de opciones para el menu 
            opciones3 = ["SELECCIONA UNA OPCION", "PENDIENTE", "EN PROCESO", "FINALIZADA", "ANULADA"]

            # Crear el menu desplegable con st.selectbox()
            seleccion3 = st.selectbox("Selecciona una opcion del menu: ", opciones3)

            if seleccion3 == "PENDIENTE":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\01. Pendiente"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()


            elif seleccion3 == "EN PROCESO":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\02. En proceso"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "FINALIZADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\03. Finalizada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

            elif seleccion3 == "ANULADA":
                #import streamlit as st
                from datetime import datetime

                def main():
                    st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                    # Sintaxis predeterminada
                    sintaxis_predeterminada = '{LF:LOOKIN="PolyfilesBO\\02. Ordenes de Trabajo\\03. OTRE - Recojo\\02. Solicitante Externo\\04. Anulada"}'

                    # Secci√≥n para seleccionar una fecha usando un calendario
                    st.markdown("<h4>Seleccione una fecha </h4>", unsafe_allow_html=True)
                    fecha_input = st.date_input("Fecha de creaci√≥n:", value=datetime.today())

                    # Variable para almacenar la sintaxis generada
                    sintaxis_final = ""

                    # Generar la sintaxis de b√∫squeda
                    if st.button("Generar Sintaxis"):
                        # Formatear la fecha seleccionada
                        fecha_formateada = fecha_input.strftime("%d/%m/%Y")
                        sintaxis_final = f"{sintaxis_predeterminada} & {{LF:Created=\"{fecha_formateada}\"}}"
                            
                        # Mostrar la sintaxis generada
                        st.success("Sintaxis de b√∫squeda generada:")
                        st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                if __name__ == "__main__":
                    main()

    def OTRE_POR_WORKORDERS():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de WOs </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        wo_numbers = content.splitlines()  # Suponiendo que cada WO est√° en una l√≠nea

                        # Crear la parte de la sintaxis para los Nros de WO
                        wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro de WO]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{wo_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                    # ==============================
                    # üîΩ Bot√≥n para descargar archivo
                    # ==============================
                    st.download_button(
                        label="üì• Descargar archivo",
                        data=sintaxis_final,
                        file_name=f"sintaxis_{datetime.now().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )

                        
        if __name__ == "__main__":
            main()

    def OTRE_POR_SOLICITUDES():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de Solicitudes </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        wo_numbers = content.splitlines()  # Suponiendo que cada WO est√° en una l√≠nea

                        # Crear la parte de la sintaxis para los Nros de WO
                        wo_sintaxis = " | ".join([f'\n{{[Orden de Trabajo]:[Nro Solicitud]="{wo.strip()}"}}' for wo in wo_numbers if wo.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{wo_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                    # ==============================
                    # üîΩ Bot√≥n para descargar archivo
                    # ==============================
                    st.download_button(
                        label="üì• Descargar archivo",
                        data=sintaxis_final,
                        file_name=f"sintaxis_{datetime.now().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )

                        
        if __name__ == "__main__":
            main()
 
    def OTRE_SOLICITUDES():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Sintaxis de Busqueda </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Nros de SOLICITUDES </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO est√° en una l√≠nea

                        # Crear la parte de la sintaxis para los Nros de WO
                        solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])

                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                    # ==============================
                    # üîΩ Bot√≥n para descargar archivo
                    # ==============================
                    st.download_button(
                        label="üì• Descargar archivo",
                        data=sintaxis_final,
                        file_name=f"sintaxis_{datetime.now().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )  

        if __name__ == "__main__":
            main()

    def COD_BASE_DE_DATO():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Codigos para descargar BD </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Codigos Poly </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO est√° en una l√≠nea

                        # Crear la parte de la sintaxis para los Nros de WO
                        solicitud_sintaxis = " OR ".join([f'\n Filefolder^ContainerCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                    # ==============================
                    # üîΩ Bot√≥n para descargar archivo
                    # ==============================
                    st.download_button(
                        label="üì• Descargar archivo",
                        data=sintaxis_final,
                        file_name=f"sintaxis_{datetime.now().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )

        if __name__ == "__main__":
            main()

    def CONTENIDO_PALLETS():
        #import streamlit as st
        from datetime import datetime

        def main():
            col1, col2 = st.columns((2))
            with col1:

                st.markdown("<h3>Generador de Piso Pallets para ver contenido </h3>", unsafe_allow_html=True)

                # Cargar archivo de texto
                st.markdown("<h4>Selecciona el archivo de texto con Locaciones de Pallets </h4>", unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Selecciona un archivo .txt", type="txt")

                # Variable para almacenar la sintaxis generada
                sintaxis_final = ""

                # Generar la sintaxis de b√∫squeda
                if st.button("Generar Sintaxis"):
                        
                    # Leer el archivo y extraer los Nros de WO
                    if uploaded_file is not None:
                        content = uploaded_file.read().decode("utf-8")
                        solicitud_numbers = content.splitlines()  # Suponiendo que cada WO est√° en una l√≠nea

                        # Crear la parte de la sintaxis para los Nros de WO
                        #solicitud_sintaxis = " | ".join([f'\n{{[Solicitud]:[Nro Solicitud]="{solicitud.strip()}"}}' for solicitud in solicitud_numbers if solicitud.strip()])
                        solicitud_sintaxis = " OR ".join([f'\n Container^LocationCode = "{solicitud.strip()}"' for solicitud in solicitud_numbers if solicitud.strip()])
                        # Construir la sintaxis final
                        sintaxis_final = f"{solicitud_sintaxis}"

            with col2:

                # Mostrar la sintaxis generada
                if sintaxis_final:
                    st.success("Sintaxis de b√∫squeda generada:")
                    st.code(sintaxis_final)  # Muestra la sintaxis en un bloque de c√≥digo

                    # ==============================
                    # üîΩ Bot√≥n para descargar archivo
                    # ==============================
                    st.download_button(
                        label="üì• Descargar archivo",
                        data=sintaxis_final,
                        file_name=f"sintaxis_{datetime.now().strftime('%Y%m%d')}.txt",
                        mime="text/plain"
                    )

        if __name__ == "__main__":
                main()

    def CODIGOS_DE_6_DIGITOS():
        import pandas as pd
        from io import StringIO

        st.title("üìÑ Completar c√≥digos a 6 d√≠gitos con ceros")

        st.markdown("""
        Sube un archivo `.txt` que contenga un c√≥digo por l√≠nea.  
        Este script completar√° cada c√≥digo a **6 d√≠gitos** rellenando con ceros a la izquierda.
        """)

        uploaded_file = st.file_uploader("üì§ Cargar archivo TXT", type=["txt"])

        if uploaded_file is not None:
            # Leer el archivo como texto
            stringio = StringIO(uploaded_file.getvalue().decode("utf-8"))
            lines = stringio.readlines()

            # Eliminar saltos de l√≠nea y espacios extra
            codes = [line.strip() for line in lines if line.strip()]

            # Completar con ceros a la izquierda
            completed_codes = [code.zfill(6) for code in codes]

            # Mostrar los primeros 5 c√≥digos
            st.subheader("üîç Vista previa (primeras 5 filas):")
            preview = completed_codes[:5]
            df_preview = pd.DataFrame(preview, columns=["C√≥digo Completado"])
            st.dataframe(df_preview, use_container_width=True)

            # Mostrar todos los resultados (colapsable)
            with st.expander("üìã Ver todos los c√≥digos completados"):
                st.code("\n".join(completed_codes), language='text')

            # Campo para que el usuario ingrese el nombre del archivo de descarga
            filename = st.text_input(
                "‚úèÔ∏è Nombre del archivo de salida (sin extensi√≥n):",
                value="codigos_completados"
            )

            # Bot√≥n de descarga
            output_text = "\n".join(completed_codes)
            if filename.strip():
                st.download_button(
                    label="‚¨áÔ∏è Descargar archivo procesado",
                    data=output_text,
                    file_name=f"{filename.strip()}.txt",
                    mime="text/plain"
                )
        else:
            st.info("Por favor, sube un archivo `.txt` para procesarlo.")

            
        # --------------------- EJECUTAR SEG√öN MEN√ö ---------------------
        
    if menu == "üß± OTPM - MATERIALES":
        OTPM_MATERIALES()
    elif menu == "üóÉÔ∏è OTEX - EXTRACIONES":
        OTEX_EXTRACIONES()
    elif menu == "üì§ OTEV - ENVIO":
        OTEV_ENVIO()
    elif menu == "üì• OTRE - RECOJO":
        OTRE_RECOJO()
    elif menu == "‚úâÔ∏è OTRE - POR_WORKORDERS":
        OTRE_POR_WORKORDERS()
    elif menu == "üßæ OTRE - POR_SOLICITUDES":
        OTRE_POR_SOLICITUDES()
    elif menu == "üìù OTRE - SOLICITUDES":
        OTRE_SOLICITUDES()
    elif menu == "üß† COD - BASE_DE_DATO":
        COD_BASE_DE_DATO()
    elif menu == "üé® CONTENIDO_PALLETS":
        CONTENIDO_PALLETS()  
    elif menu == "üîñ CODIGOS DE 6 DIGITOS":
        CODIGOS_DE_6_DIGITOS()


elif proyecto == "DUPLICAR FILAS":
    import streamlit as st
    import pandas as pd
    from io import BytesIO

    st.subheader("üìë DUPLICAR FILAS DE ARCHIVO EXCEL")

    # Cargar archivo Excel
    uploaded_file = st.file_uploader("Elige un archivo Excel", type=["xlsx"])

    if uploaded_file is not None:
        # Leer el archivo Excel
        df = pd.read_excel(uploaded_file)

        # Mostrar el DataFrame original
        st.write("DataFrame Original:")
        st.dataframe(df)

        # Duplicar filas
        if st.button("Duplicar Filas"):
            # Crear un nuevo DataFrame duplicando cada fila
            duplicated_rows = pd.DataFrame(columns=df.columns)
                
            for index, row in df.iterrows():
                duplicated_rows = pd.concat([duplicated_rows, pd.DataFrame([row])], ignore_index=True)
                duplicated_rows = pd.concat([duplicated_rows, pd.DataFrame([row])], ignore_index=True)

            st.write("DataFrame Duplicado:")
            st.dataframe(duplicated_rows)

            # Opci√≥n para descargar el archivo duplicado como Excel
            def to_excel(df):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Duplicado')
                return output.getvalue()

            excel_data = to_excel(duplicated_rows)
            st.download_button("Descargar archivo duplicado como Excel", excel_data, "duplicado.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


elif proyecto == "JUEGA CON DINO":

    import streamlit as st

    st.subheader("ü¶ñ Dinosaurio Saltar√≠n ‚òÅÔ∏è")

    # Bot√≥n de reinicio
    if st.button("üîÑ Reiniciar Juego"):
        st.rerun()

    game_code = """
    <style>
    html, body {
        margin: 0;
        padding: 0;
        height: 100%;
        background-color: black;
    }

    #gameContainer {
        width: 100%;
        height: 100%;
        background-color: black;
    }

    canvas {
        display: block;
        width: 100%;
        height: 100%;
        background-color: #cce7ff;  /* D√≠a por defecto */
        border: 2px solid #333;
    }
    </style>

    <div id="gameContainer">
    <canvas id="gameCanvas"></canvas>
    </div>

    <script>
    const canvas = document.getElementById("gameCanvas");
    const ctx = canvas.getContext("2d");

    function resizeCanvas() {
        canvas.width = canvas.clientWidth;
        canvas.height = canvas.clientHeight;
    }
    resizeCanvas();
    window.addEventListener("resize", resizeCanvas);

    ctx.textBaseline = "bottom";

    let groundLevel = canvas.height - 20;

    let dino = {
        x: 50, y: groundLevel - 40, width: 40, height: 40,
        dy: 0, gravity: 1, jumpPower: -15, grounded: true
    };

    let obstacles = [
        { x: 800, width: 20, height: 40, type: "smallCactus", speed: 6 },
        { x: 1200, width: 30, height: 60, type: "bigCactus", speed: 6 },
        { x: 1600, width: 30, height: 30, type: "bird", speed: 8 }
    ];

    // Ajustar posici√≥n inicial
    obstacles.forEach(o => {
        if (o.type === "bird") {
            o.y = groundLevel - o.height - 60;
        } else {
            o.y = groundLevel - o.height;
        }
    });

    // Nubes decorativas ‚òÅÔ∏è
    let clouds = [
        { x: 100, y: 50, speed: 1 },
        { x: 400, y: 80, speed: 0.8 },
        { x: 700, y: 60, speed: 1.2 }
    ];

    function drawClouds() {
        ctx.font = "30px Arial";
        clouds.forEach(c => {
            ctx.fillText("‚òÅÔ∏è", c.x, c.y);
        });
    }

    function updateClouds() {
        clouds.forEach(c => {
            c.x -= c.speed;
            if (c.x < -50) {
                c.x = canvas.width + Math.random() * 200;
                c.y = 30 + Math.random() * 60;
            }
        });
    }

    // D√≠a y noche üåûüåô
    let isDay = true;

    function updateBackgroundColor() {
        if (score % 20 === 0 && score !== 0 && score !== lastToggleScore) {
            isDay = !isDay;
            lastToggleScore = score;
        }

        if (isDay) {
            canvas.style.backgroundColor = "#cce7ff";  // D√≠a
        } else {
            canvas.style.backgroundColor = "#2c3e50";  // Noche
        }
    }

    let gameOver = false;
    let score = 0;
    let lastToggleScore = -1;

    function drawBackground() {
        ctx.fillStyle = "#8B4513";
        ctx.fillRect(0, canvas.height - 20, canvas.width, 20);
    }

    function drawDino() {
        ctx.font = "40px Arial";
        ctx.fillText("ü¶ñ", dino.x, dino.y + dino.height);
    }

    function drawObstacles() {
        ctx.font = "40px Arial";
        obstacles.forEach(o => {
            if (o.type === "smallCactus" || o.type === "bigCactus") {
                ctx.fillText("üåµ", o.x, o.y + o.height);
            } else if (o.type === "bird") {
                ctx.fillText("üê¶", o.x, o.y + o.height);
            }
        });
    }

    function update() {
        if (gameOver) return;

        ctx.clearRect(0, 0, canvas.width, canvas.height);

        updateBackgroundColor();  // Cambiar d√≠a/noche
        updateClouds();
        drawBackground();
        drawClouds();

        groundLevel = canvas.height - 20;

        // Dino saltando
        dino.y += dino.dy;
        if (dino.y + dino.height < groundLevel) {
            dino.dy += dino.gravity;
            dino.grounded = false;
        } else {
            dino.y = groundLevel - dino.height;
            dino.dy = 0;
            dino.grounded = true;
        }

        // Obst√°culos
        obstacles.forEach(o => {
            o.x -= o.speed;
            if (o.x + o.width < 0) {
                o.x = canvas.width + Math.random() * 400;

                if (o.type === "bird") {
                    o.y = groundLevel - o.height - 60;
                } else {
                    o.y = groundLevel - o.height;
                }

                score++;
                if (score % 5 === 0) o.speed += 1;
            }

            // Colisi√≥n
            if (
                dino.x < o.x + o.width &&
                dino.x + dino.width > o.x &&
                dino.y < o.y + o.height &&
                dino.y + dino.height > o.y
            ) {
                gameOver = true;
                alert("üíÄ Game Over! Fuck you Puntuaci√≥n final: " + score);
            }
        });

        drawDino();
        drawObstacles();

        ctx.fillStyle = isDay ? "black" : "white";
        ctx.font = "20px Arial";
        ctx.fillText("Score: " + score, 10, 30);

        requestAnimationFrame(update);
    }

    window.addEventListener("keydown", function(e) {
        if ((e.code === "Space" || e.code === "ArrowUp") && dino.grounded) {
            dino.dy = dino.jumpPower;
        }
    });

    update();
    </script>
    """

    st.components.v1.html(game_code, height=450)


else:
    st.title("üëã Bienvenido al Sistema de Almac√©n")

    st.write("Seleccione una opci√≥n del men√∫ lateral")

